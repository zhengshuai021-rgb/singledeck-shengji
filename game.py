#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一副牌升级游戏模拟器
根据 PRD 实现完整游戏逻辑：发牌 → 定主 → 埋底 → 捡主 → 出牌 → 结算 → 过7

修改时间：2026-05-27 17:08
修改：①过7改为完整循环（7→8→...→6→7）②移除对子/拖拉机（一副牌无同花色对子）

创建时间：2026-05-27
作者：Kami 🐱
"""

import random
import os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# ==================== 常量 ====================

SUITS = ['♠', '♥', '♣', '♦']
RANKS = ['2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A']
SCORE_RANKS = {'5', '10', 'K'}
SCORE_VALUES = {'5': 5, '10': 10, 'K': 10}
RANK_ORDER = {rank: idx for idx, rank in enumerate(RANKS)}
SUIT_CN = {'♠': '黑桃', '♥': '红桃', '♣': '草花', '♦': '方块'}

# 升级序列：7→8→9→10→J→Q→K→A→2→3→4→5→6→7(完成过7)
# 内部：7=0, 8=1, ..., A=7, 2=8, 3=9, 4=10, 5=11, 6=12
LEVEL_CYCLE = ['7', '8', '9', '10', 'J', 'Q', 'K', 'A', '2', '3', '4', '5', '6']
LEVEL_CYCLE_LEN = len(LEVEL_CYCLE)  # 13

def level_idx(lvl):
    """等级→索引（7→0, 8→1, ..., 6→12）"""
    if lvl in ('2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A'):
        return LEVEL_CYCLE.index(lvl)
    return 0

def level_up(lvl, steps=1):
    """升级：返回新等级"""
    idx = level_idx(lvl)
    return LEVEL_CYCLE[(idx + steps) % LEVEL_CYCLE_LEN]

# ==================== 牌类 ====================

class Card:
    def __init__(self, suit, rank):
        self.suit = suit
        self.rank = rank
    def __repr__(self):
        if self.rank in ('大王', '小王'): return self.rank
        return f"{self.suit}{self.rank}"
    def __eq__(self, other):
        return isinstance(other, Card) and self.suit == other.suit and self.rank == other.rank
    def __hash__(self):
        return hash((self.suit, self.rank))

def create_deck():
    deck = []
    for suit in SUITS:
        for rank in RANKS:
            deck.append(Card(suit, rank))
    deck.append(Card('王', '大王'))
    deck.append(Card('王', '小王'))
    return deck

def cards_str(cards):
    return ' '.join(str(c) for c in cards) if cards else '无'

# ==================== 牌力判定 ====================

def is_main(card, level, trump_suit):
    """判断是否主牌"""
    if card.rank in ('大王', '小王'): return True
    if card.rank == '3' and card.suit == '♥': return True
    if card.rank == '2': return True
    if card.rank == level: return True
    return False

def card_power(card, level, trump_suit):
    """(group, value): 5=大王 4=小王 3=级牌 2=常主 1=主花色 0=副牌"""
    if card.rank == '大王': return (5, 100)
    if card.rank == '小王': return (4, 100)
    if card.rank == level: return (3, RANK_ORDER[card.rank])
    if card.rank == '3' and card.suit == '♥': return (2, 100)
    if card.rank == '2': return (2, 90)
    if card.suit == trump_suit: return (1, RANK_ORDER[card.rank])
    return (0, RANK_ORDER[card.rank])

def cp(card, level, trump_suit):
    """综合牌力值"""
    g, v = card_power(card, level, trump_suit)
    return g * 1000 + v

def compare_cards(c1, c2, level, trump_suit, lead_suit):
    """1=c1大, -1=c2大 (单张)"""
    c1_lead = c1.suit == lead_suit
    c2_lead = c2.suit == lead_suit
    c1_main = is_main(c1, level, trump_suit)
    c2_main = is_main(c2, level, trump_suit)

    if c1_lead and c2_lead:
        p1, p2 = card_power(c1, level, trump_suit), card_power(c2, level, trump_suit)
        if p1[0] != p2[0]: return 1 if p1[0] > p2[0] else -1
        return 1 if p1[1] > p2[1] else (-1 if p1[1] < p2[1] else 1)
    if c1_lead and not c2_lead:
        return -1 if c2_main else 1
    if not c1_lead and c2_lead:
        return 1 if c1_main else -1
    if c1_main and c2_main:
        p1, p2 = card_power(c1, level, trump_suit), card_power(c2, level, trump_suit)
        if p1[0] != p2[0]: return 1 if p1[0] > p2[0] else -1
        return 1 if p1[1] > p2[1] else (-1 if p1[1] < p2[1] else 1)
    if c1_main: return 1
    if c2_main: return -1
    return 1  # 都是非主垫牌，先出算大

# ==================== 牌型系统 ====================

def group_by_rank(cards):
    """按牌值分组: {rank: [cards]}"""
    d = defaultdict(list)
    for c in cards: d[c.rank].append(c)
    return d

def find_510k(hand, suit):
    """检查手牌中是否有同花色 5+10+K 组合，返回 [Card, ...] 或 None"""
    suit_cards = [c for c in hand if c.suit == suit and c.rank in ('5', '10', 'K')]
    ranks_present = {c.rank for c in suit_cards}
    if ranks_present == {'5', '10', 'K'}:
        return [c for c in suit_cards]
    return None

def find_hongs(hand):
    """找出手牌中所有"轰"（4张同点数），返回 [(rank, [cards]), ...]"""
    groups = group_by_rank(hand)
    return [(rank, cards) for rank, cards in groups.items() if len(cards) == 4]

def find_zhas(hand):
    """找出手牌中所有"炸"（♠A+3张同点数），返回 [(rank, [cards]), ...]
    炸不可首出，仅能跟轰时垫出
    """
    groups = group_by_rank(hand)
    sa_list = [c for c in groups.get('A', []) if c.suit == '♠']
    if not sa_list:
        return []
    result = []
    for rank, cards in groups.items():
        if rank == 'A':
            continue
        if len(cards) >= 3:
            result.append((rank, [sa_list[0]] + cards[:3]))
    return result

# 牌型层级: 5=大王+小王+♥3  4=主510K  3=副510K  2=轰  1=炸  0=单张
def pattern_hierarchy(pattern, cards, level, trump_suit):
    """返回牌型层级（越大越强）和排序值"""
    if pattern == '510k':
        is_main_510k = all(is_main(c, level, trump_suit) for c in cards)
        return (4 if is_main_510k else 3, 0)
    if pattern == 'hong':
        rank = cards[0].rank
        return (2, RANK_ORDER.get(rank, 0))
    if pattern == 'zha':
        r = cards[1].rank if cards[0].rank == 'A' else cards[0].rank
        return (1, RANK_ORDER.get(r, 0))
    return (0, 0)

def compare_trick_patterns(pattern_a, cards_a, pattern_b, cards_b, level, trump_suit, lead_suit):
    """牌型级比较：1=A大, -1=B大, 0=平
    规则：
      - 不同牌型层级 → 层级高的直接赢
      - 同牌型 → 510K比同花色内最高牌, 轰比牌值, 炸比牌值
      - 都是单张 → fallback compare_cards
      - 层级不同：单张无论大小都输给510K/轰/炸
    """
    ha = pattern_hierarchy(pattern_a, cards_a, level, trump_suit)
    hb = pattern_hierarchy(pattern_b, cards_b, level, trump_suit)

    # 层级不同 → 直接判
    if ha[0] != hb[0]:
        return 1 if ha[0] > hb[0] else -1

    # 同层级，内部比较
    if ha[0] == 0:  # 都是单张
        return compare_cards(cards_a[0], cards_b[0], level, trump_suit, lead_suit)
    if ha[0] in (4, 3):  # 都是510K，同花色/主色内比较
        # 510K 内部最大牌值
        max_a = max(RANK_ORDER.get(c.rank, 0) for c in cards_a)
        max_b = max(RANK_ORDER.get(c.rank, 0) for c in cards_b)
        return 1 if max_a > max_b else (-1 if max_a < max_b else 1)
    if ha[0] == 2:  # 都是轰，比牌值
        r_a = cards_a[0].rank
        r_b = cards_b[0].rank
        return 1 if RANK_ORDER.get(r_a, 0) > RANK_ORDER.get(r_b, 0) else -1
    if ha[0] == 1:  # 都是炸，比内部牌值
        r_a = cards_a[1].rank if cards_a[0].rank == 'A' else cards_a[0].rank
        r_b = cards_b[1].rank if cards_b[0].rank == 'A' else cards_b[0].rank
        return 1 if RANK_ORDER.get(r_a, 0) > RANK_ORDER.get(r_b, 0) else -1
    return 1

def max_card_in_trick(played, level, trump_suit, lead_suit):
    """返回一圈中最大的 (pid, card)"""
    flat = []
    for item in played:
        pid = item[0]
        cards_or_card = item[1]
        if isinstance(cards_or_card, list):
            for card in cards_or_card:
                flat.append((pid, card))
        else:
            flat.append((pid, cards_or_card))

    if not flat:
        return None, None

    best_pid, best_card = flat[0]
    for pid, card in flat[1:]:
        if compare_cards(card, best_card, level, trump_suit, lead_suit) == 1:
            best_pid, best_card = pid, card
    return best_pid, best_card

# ==================== 机器人 ====================

class Bot:
    def __init__(self, pid, hand, side, level, trump_suit):
        self.pid = pid
        self.hand = list(hand)
        self.side = side
        self.level = level
        self.trump_suit = trump_suit

    def _non_score_offsuit(self):
        """非分牌且非主牌的牌"""
        return [c for c in self.hand if not is_main(c, self.level, self.trump_suit)
                and c.rank not in SCORE_RANKS]

    def _all_main(self):
        """所有主牌"""
        return [c for c in self.hand if is_main(c, self.level, self.trump_suit)]

    def _detect_pattern(self, cards):
        """检测牌型：'510k' | 'hong' | 'zha' | 'single'"""
        if len(cards) == 1:
            return 'single'
        if len(cards) == 3:
            if {c.rank for c in cards} == {'5', '10', 'K'}:
                return '510k'
        if len(cards) == 4:
            ranks = [c.rank for c in cards]
            if len(set(ranks)) == 1:
                return 'hong'
            if len(set(ranks)) == 2 and 'A' in ranks and ranks.count('A') == 1:
                return 'zha'
        return 'single'

    def lead(self):
        """首出：优先副牌轰 > 副510K > 单张非分非主"""
        level, ts = self.level, self.trump_suit

        if not self.hand:
            return []

        # 1. 副牌轰
        for rank, cards in find_hongs(self.hand):
            if not all(is_main(c, level, ts) for c in cards):
                for c in cards: self.hand.remove(c)
                return cards

        # 2. 副牌510K
        for suit in SUITS:
            if suit == ts: continue
            cards = find_510k(self.hand, suit)
            if cards:
                for c in cards: self.hand.remove(c)
                return cards

        # 3. 单张：优先非分非主小牌
        safe = self._non_score_offsuit()
        if safe:
            safe.sort(key=lambda c: RANK_ORDER.get(c.rank, 0))
            card = safe[0]
        else:
            ns = [c for c in self.hand if c.rank not in SCORE_RANKS]
            card = min(ns if ns else self.hand, key=lambda c: RANK_ORDER.get(c.rank, 0))
        self.hand.remove(card)
        return [card]

    def follow(self, lead_suit, played_so_far, is_last_trick=False):
        """跟牌：数量必须与首出相等，不够用其他牌凑"""
        level, ts = self.level, self.trump_suit

        if not self.hand:
            return []

        first_cards = played_so_far[0][1] if played_so_far else []
        if not first_cards:
            return self._discard_or_trump(played_so_far, is_last_trick, need=1)

        # 手牌总数不足时，全部打出
        need_count = len(first_cards)
        if len(self.hand) <= need_count:
            out = list(self.hand)
            self.hand.clear()
            return out

        lead_pattern = self._detect_pattern(first_cards)

        # ====== 5·10·K ======
        if lead_pattern == '510k':
            all_main = self._all_main()
            main_510k = find_510k(self.hand, ts)
            supreme = [c for c in self.hand if c.rank in ('大王', '小王')
                       or (c.rank == '3' and c.suit == '♥')]
            if len(supreme) >= 3 and main_510k is None:
                out = supreme[:3]
                for c in out: self.hand.remove(c)
                return out
            if main_510k:
                for c in main_510k: self.hand.remove(c)
                return main_510k
            # 没有 510K/王炸 → 用主牌凑够数量，不够用副牌补
            if all_main:
                all_main.sort(key=lambda c: cp(c, level, ts))
                out = all_main[:need_count]
                for c in out: self.hand.remove(c)
                # 主牌不够，用副牌补齐
                if len(out) < need_count:
                    off = [c for c in self.hand if not is_main(c, level, ts)]
                    out += off[:need_count - len(out)]
                    for c in off[:need_count - len(out)]: self.hand.remove(c)
                return out
            # 用副牌凑够数量
            off = [c for c in self.hand if not is_main(c, level, ts)]
            out = off[:need_count]
            for c in out: self.hand.remove(c)
            return out

        # ====== 轰 ======
        if lead_pattern == 'hong':
            lead_rank = first_cards[0].rank
            lead_hong_main = all(is_main(c, level, ts) for c in first_cards)

            for rank, cards in find_hongs(self.hand):
                card_main = all(is_main(c, level, ts) for c in cards)
                can_beat = False
                if card_main and not lead_hong_main:
                    can_beat = True
                elif card_main == lead_hong_main:
                    can_beat = RANK_ORDER.get(rank, 0) > RANK_ORDER.get(lead_rank, 0)
                if can_beat:
                    for c in cards: self.hand.remove(c)
                    return cards

            # 垫炸
            zhas = find_zhas(self.hand)
            if zhas:
                rank, cards = zhas[0]
                for c in cards: self.hand.remove(c)
                return cards

            # 出主牌
            all_main = self._all_main()
            if len(all_main) >= len(first_cards):
                all_main.sort(key=lambda c: cp(c, level, ts))
                out = all_main[:len(first_cards)]
                for c in out: self.hand.remove(c)
                return out

            out = list(self.hand[:len(first_cards)])
            for c in out: self.hand.remove(c)
            return out

        # ====== 炸 ======
        if lead_pattern == 'zha':
            for rank, cards in find_hongs(self.hand):
                for c in cards: self.hand.remove(c)
                return cards
            return self._discard_or_trump(played_so_far, is_last_trick, need=len(first_cards))

        # ====== 单张 ======
        same = [c for c in self.hand if c.suit == lead_suit]
        has_score = any(c.rank in SCORE_RANKS
                        for _, cl in played_so_far for c in cl)
        need_count = len(first_cards)

        if same:
            # 有首出花色：同花色出够数量，不够用其他牌补
            if has_score:
                best_card = self._find_best_to_win(same, played_so_far, level, ts, lead_suit)
                if best_card:
                    same.remove(best_card)
                # 凑齐数量：先同花色，再其他牌
                out = [best_card] + same[:need_count - 1] if best_card else same[:need_count]
            else:
                ns = [c for c in same if c.rank not in SCORE_RANKS]
                pool = ns if ns else same
                out = pool[:need_count]
            for c in out: self.hand.remove(c)
            return out
        else:
            return self._discard_or_trump(played_so_far, is_last_trick, need=need_count)

    def _discard_or_trump(self, played_so_far, is_last_trick, need=1):
        level, ts = self.level, self.trump_suit

        lead_suit = None
        for _, cl in played_so_far:
            if cl:
                lead_suit = cl[0].suit
                break
        if lead_suit is None:
            lead_suit = '♠'

        has_score = any(c.rank in SCORE_RANKS
                        for _, cl in played_so_far for c in cl)

        all_main = self._all_main()
        out = None

        if has_score:
            best_pid, best_card = max_card_in_trick(played_so_far, level, ts, lead_suit)
            if best_card:
                best_p = cp(best_card, level, ts)

                if self.side == 'dealer':
                    can_win = [c for c in all_main if cp(c, level, ts) > best_p]
                    if can_win:
                        can_win.sort(key=lambda c: cp(c, level, ts))
                        out = can_win[:need]
                    if out is None:
                        main = [c for c in all_main
                                if c.rank not in ('大王', '小王')
                                and not (c.rank == '3' and c.suit == '♥')]
                        if main:
                            main.sort(key=lambda c: RANK_ORDER.get(c.rank, 0))
                            out = main[:need]
                else:
                    can_win = [c for c in all_main if cp(c, level, ts) > best_p]
                    if can_win:
                        can_win.sort(key=lambda c: cp(c, level, ts))
                        out = can_win[:need]

        if out is None:
            off = [c for c in self.hand if not is_main(c, level, ts)]
            off.sort(key=lambda c: RANK_ORDER.get(c.rank, 0))
            out = off[:need]
            if not out:
                out = list(self.hand[:need])

        for c in out:
            if c in self.hand:
                self.hand.remove(c)
        if len(out) < need:
            extra = list(self.hand[:need - len(out)])
            for c in extra:
                self.hand.remove(c)
            out += extra
        return out

    def _find_best_to_win(self, same_suit_cards, played_so_far, level, ts, lead_suit):
        """在同花色中找能赢的最小牌"""
        if not played_so_far: return None
        best_pid, best_card = max_card_in_trick(played_so_far, level, ts, lead_suit)
        if best_card is None: return None
        best_p = cp(best_card, level, ts)
        can_win = [c for c in same_suit_cards if cp(c, level, ts) > best_p]
        if can_win:
            can_win.sort(key=lambda c: cp(c, level, ts))
            return can_win[0]
        return None

    def select_for_bottom(self, count):
        """选牌埋底/弃回：优先非分非主小牌"""
        def pri(c):
            return (int(is_main(c, self.level, self.trump_suit)),
                    int(c.rank in SCORE_RANKS),
                    RANK_ORDER.get(c.rank, 0))
        s = sorted(self.hand, key=pri)
        sel = s[:count]
        for c in sel: self.hand.remove(c)
        return sel

# ==================== 游戏记录 ====================

class RoundRecord:
    def __init__(self, rnd, dealer_pid, def_lvl, att_lvl):
        self.rnd = rnd
        self.dealer_pid = dealer_pid
        self.defender_level = def_lvl
        self.attacker_level = att_lvl
        self.level = def_lvl
        self.initial_hands = {}
        self.initial_bottom = []
        self.bright_pid = None; self.bright_card = None
        self.concealed_pid = None; self.concealed_card = None
        self.trump_suit = None
        self.trump_method = None
        self.buried_cards = []
        self.bottom_after_bury = []
        self.picked_from_bottom = []
        self.discarded_to_bottom = []
        self.bottom_after_pick = []
        self.tricks = []
        self.attacker_score = 0
        self.last_trick_winner_pid = None
        self.last_trick_winner_side = None
        self.last_trick_card = None
        self.base_up_att = 0
        self.bonus_up = 0
        self.final_up_att = 0
        self.final_up_def = 0
        self.result = ''
        self.result_title = ''
        self.logs = []
        self.dealer_team = []
        self.attacker_team = []
        self.game_over_check = False
        # 轮次相关
        self.round_ended = False
        self.round_winner = None

    def log(self, msg): self.logs.append(msg)

# ==================== 游戏引擎 ====================

class Game:
    def __init__(self, total_rounds=None, max_games=None):
        # 庄家方级牌，起始为7，过7需要完整走一圈回到7
        self.defender_level = '7'
        # 抓分方级牌，起始为7
        self.attacker_level = '7'
        self.dealer_pid = random.randint(0, 3)
        self.records = []
        self.game_over = False
        self.winner = None
        self.rnd = 0
        # 按队伍追踪等级（不受庄权交换影响）
        # 队伍A = 初始庄家方（玩家0+2），队伍B = 初始抓分方（玩家1+3）
        self.team_a_level = '7'  # 队伍A等级
        self.team_b_level = '7'  # 队伍B等级
        self.team_a_cumulative_steps = 0  # 队伍A累计升级步数（跨局）
        self.team_b_cumulative_steps = 0  # 队伍B累计升级步数（跨局）
        self.defending_team = None  # 守庄方（None=None / 'A' / 'B'）
        # 轮次追踪
        self.total_rounds = total_rounds  # 总轮数（None=不限制）
        self.max_games = max_games if max_games is not None else 200  # 总局数上限
        self.current_round = 1  # 当前第几轮
        self.round_starts_at = 1  # 本轮从第几局开始
        self.round_records = []  # [{round, start_rnd, end_rnd, winner, games_count}, ...]

    def run(self):
        while not self.game_over and (self.max_games <= 0 or self.rnd < self.max_games):
            self.rnd += 1
            rec = self._play_round()
            self.records.append(rec)
            # 检查轮次是否结束
            if rec.round_ended:
                self.round_records.append({
                    'round': self.current_round,
                    'start_rnd': self.round_starts_at,
                    'end_rnd': self.rnd,
                    'winner': rec.round_winner or '—',
                    'games_count': self.rnd - self.round_starts_at + 1,
                })
                self.current_round += 1
                self.round_starts_at = self.rnd + 1
            # 检查是否达到总轮数
            if self.total_rounds and self.current_round > self.total_rounds:
                self.game_over = True
        return self.records

    def _play_round(self):
        rec = RoundRecord(self.rnd, self.dealer_pid, self.defender_level, self.attacker_level)
        dt = [self.dealer_pid, (self.dealer_pid + 2) % 4]
        at = [(self.dealer_pid + 1) % 4, (self.dealer_pid + 3) % 4]
        rec.dealer_team = dt
        rec.attacker_team = at

        rec.log(f"=== 第{self.rnd}局 ===")
        rec.log(f"庄家方级牌={self.defender_level}, 抓分方级牌={self.attacker_level}, 庄家=玩家{self.dealer_pid+1}")
        rec.log(f"庄家阵营: 玩{dt[0]+1}、玩{dt[1]+1} | 抓分阵营: 玩{at[0]+1}、玩{at[1]+1}")

        hands, bottom = self._deal(rec)

        for stop_count in range(2):
            should_stop = False
            stop_pid = None
            for pid in at:
                if not any(c.rank in SCORE_RANKS for c in hands[pid]):
                    if random.random() < 0.5:
                        should_stop = True
                        stop_pid = pid
                        break
            if should_stop:
                rec.log(f"【停级】闲家玩{stop_pid+1}无分牌，停级第{stop_count+1}次，重新发牌")
                deck = create_deck()
                random.shuffle(deck)
                hands = [[] for _ in range(4)]
                bottom = []
                for i, card in enumerate(deck):
                    (hands[i % 4] if i < 48 else bottom).append(card)
                rec.initial_hands = {p: list(h) for p, h in enumerate(hands)}
                rec.initial_bottom = list(bottom)
                rec.log(f"重新发牌 | 底牌: {cards_str(bottom)}")
            else:
                break

        self._determine_trump(rec, hands)

        # 关键：闷牌玩家恰好是庄家 → 闷牌自动变亮牌
        # 因为庄家就是定庄方，不存在闲家闷牌，所以不能捡主
        if (rec.trump_method == 'concealed'
                and rec.concealed_pid is not None
                and rec.concealed_pid in rec.dealer_team
                and rec.concealed_pid == rec.dealer_pid):
            rec.bright_pid = rec.concealed_pid
            rec.bright_card = rec.concealed_card
            rec.trump_suit = rec.concealed_card.suit
            rec.trump_method = 'bright'  # 改为亮牌
            rec.log(f"【闷→亮】玩{rec.concealed_pid+1} 是庄家，闷牌自动变亮牌 {rec.bright_card} → 主={SUIT_CN[rec.trump_suit]}")
            rec.concealed_pid = None
            rec.concealed_card = None

        self._bury(rec, hands)
        # 只有闲家闷牌才捡主（庄家闷已转亮牌，无捡主）
        if rec.concealed_pid is not None:
            self._pick_main(rec, hands)

        for pid in range(4):
            assert len(hands[pid]) == 12, f"玩{pid+1} 手牌 {len(hands[pid])} != 12"

        self._play_tricks(rec, hands, dt, at)
        self._settle(rec)

        rec.log(f"第{self.rnd}局结束 | 庄家方={self.defender_level} 抓分方={self.attacker_level}")
        return rec

    def _deal(self, rec):
        deck = create_deck()
        random.shuffle(deck)
        hands = [[] for _ in range(4)]
        bottom = []
        for i, card in enumerate(deck):
            (hands[i % 4] if i < 48 else bottom).append(card)
        rec.initial_hands = {p: list(h) for p, h in enumerate(hands)}
        rec.initial_bottom = list(bottom)
        rec.log(f"发牌完成 | 底牌: {cards_str(bottom)}")
        return hands, bottom

    def _determine_trump(self, rec, hands):
        dt, at = rec.dealer_team, rec.attacker_team
        lvl = rec.level

        # 未定庄阶段：全员可闷牌（有1人闷即停止）
        for pid in range(4):
            lc = [c for c in hands[pid] if c.rank == lvl and c.suit in SUITS]
            if lc and random.random() < 0.25:
                card = random.choice(lc)
                rec.concealed_pid, rec.concealed_card = pid, card
                rec.trump_method = 'concealed'
                rec.log(f"【闷牌】玩{pid+1} 闷一张级牌（花色待揭晓）")
                return

        # 定庄后：庄家亮牌与闲家闷牌互斥
        for pid in dt:
            lc = [c for c in hands[pid] if c.rank == lvl and c.suit in SUITS]
            if lc and random.random() < 0.5:
                card = random.choice(lc)
                rec.bright_pid, rec.bright_card = pid, card
                rec.trump_suit, rec.trump_method = card.suit, 'bright'
                rec.log(f"【亮牌】玩{pid+1} 亮 {card} → 主={SUIT_CN[card.suit]}")
                return

        for pid in at:
            lc = [c for c in hands[pid] if c.rank == lvl and c.suit in SUITS]
            if lc and random.random() < 0.5:
                card = random.choice(lc)
                rec.concealed_pid, rec.concealed_card = pid, card
                rec.trump_method = 'concealed'
                rec.log(f"【闷牌】玩{pid+1} 闷一张级牌（花色待揭晓）")
                return

        fc = next((c for c in rec.initial_bottom if c.suit in SUITS), None)
        rec.trump_suit = fc.suit if fc else random.choice(SUITS)
        rec.trump_method = 'bottom_card'
        rec.log(f"【底牌首张定主】{fc} → 主={SUIT_CN.get(rec.trump_suit, rec.trump_suit)}")

    def _bury(self, rec, hands):
        pid = rec.dealer_pid
        bottom = list(rec.initial_bottom)
        bot = Bot(pid, hands[pid], 'dealer', rec.level, rec.trump_suit or '')

        n = random.randint(0, 6)
        buried = bot.select_for_bottom(n)
        temp_bottom = bottom + buried

        score = sum(SCORE_VALUES.get(c.rank, 0) for c in temp_bottom)
        for _ in range(10):
            if score <= 35: break
            sc = [c for c in buried if c.rank in SCORE_RANKS]
            if not sc: break
            buried.remove(sc[0]); bot.hand.append(sc[0])
            temp_bottom = bottom + buried
            score = sum(SCORE_VALUES.get(c.rank, 0) for c in temp_bottom)

        # take_back 数量 = 当前 buried 数量（而非原始 n），确保 new_bottom = 6
        take_back = temp_bottom[:len(buried)]
        new_bottom = temp_bottom[len(buried):]
        assert len(new_bottom) == 6, f"底牌数 {len(new_bottom)} != 6 (buried={len(buried)}, temp_bottom={len(temp_bottom)})"
        bot.hand.extend(take_back)
        hands[pid] = bot.hand

        rec.buried_cards, rec.bottom_after_bury = list(buried), list(new_bottom)
        bs = sum(SCORE_VALUES.get(c.rank, 0) for c in new_bottom)
        rec.log(f"【埋底】庄家（玩{pid+1}）埋 {cards_str(buried)}, 取回 {cards_str(take_back)}")
        rec.log(f"  埋底后底牌: {cards_str(new_bottom)} (分值={bs})")

    def _pick_main(self, rec, hands):
        pid = rec.concealed_pid
        ts = rec.concealed_card.suit
        rec.trump_suit = ts
        bottom = list(rec.bottom_after_bury)
        bot = Bot(pid, hands[pid], 'attacker', rec.level, ts)

        rec.log(f"【捡主】玩{pid+1} 翻开闷牌 {rec.concealed_card} → 主={SUIT_CN[ts]}")
        picked = [c for c in bottom if is_main(c, rec.level, ts)]
        if not picked:
            rec.log(f"  底牌无主牌，跳过"); rec.bottom_after_pick = list(bottom); return

        rec.picked_from_bottom = list(picked)
        bottom_rem = [c for c in bottom if c not in picked]
        bot.hand.extend(picked)
        discarded = bot.select_for_bottom(len(picked))
        new_bottom = bottom_rem + discarded
        new_bs = sum(SCORE_VALUES.get(c.rank, 0) for c in new_bottom)
        if new_bs > 35:
            rec.log(f"  捡主后底牌分值={new_bs}>35，不可捡主")
            rec.bottom_after_pick = list(bottom)
            return
        assert len(new_bottom) == 6
        hands[pid] = list(bot.hand)
        rec.discarded_to_bottom = list(discarded)
        rec.bottom_after_pick = list(new_bottom)
        bs = sum(SCORE_VALUES.get(c.rank, 0) for c in new_bottom)
        rec.log(f"  拣出: {cards_str(picked)} | 弃回: {cards_str(discarded)}")
        rec.log(f"  捡主后底牌: {cards_str(new_bottom)} (分值={bs})")

    def _play_tricks(self, rec, hands, dt, at):
        bots = {}
        for pid in range(4):
            side = 'dealer' if pid in dt else 'attacker'
            bots[pid] = Bot(pid, hands[pid], side, rec.level, rec.trump_suit)

        leader = rec.dealer_pid
        t = 0
        max_tricks = 12

        while any(bots[pid].hand for pid in range(4)) and t < max_tricks:
            t += 1
            trick = {'num': t, 'leader': leader, 'played': [], 'winner': None,
                     'winner_side': None, 'score': 0, 'score_cards': [], 'pattern': 'single'}
            lead_suit = None
            played_so_far = []

            for pos in range(4):
                pid = (leader + pos) % 4
                if not bots[pid].hand:
                    trick['played'].append((pid, []))
                    played_so_far.append((pid, []))
                    continue
                if pos == 0:
                    card_list = bots[pid].lead()
                    lead_suit = card_list[0].suit if card_list else None
                else:
                    card_list = bots[pid].follow(lead_suit, played_so_far, is_last_trick=False)
                trick['played'].append((pid, card_list))
                played_so_far.append((pid, card_list))

            if not any(cl for _, cl in trick['played']):
                t -= 1
                continue

            first_cards = played_so_far[0][1] if played_so_far else []
            if first_cards:
                bot0 = bots[played_so_far[0][0]]
                trick['pattern'] = bot0._detect_pattern(first_cards) if bot0 else 'single'

            # 牌型级比较：找赢家
            best_pid = None
            best_pattern = None
            best_cards = None
            for pid, card_list in trick['played']:
                if not card_list:
                    continue
                p = bots[pid]._detect_pattern(card_list) if pid in bots else 'single'
                if best_pid is None:
                    best_pid, best_pattern, best_cards = pid, p, card_list
                else:
                    cmp = compare_trick_patterns(
                        p, card_list, best_pattern, best_cards,
                        rec.level, rec.trump_suit, lead_suit
                    )
                    if cmp == 1:
                        best_pid, best_pattern, best_cards = pid, p, card_list

            if best_pid is None:
                t -= 1
                continue

            trick['winner'] = best_pid
            trick['winner_side'] = 'dealer' if best_pid in dt else 'attacker'
            trick['winner_pattern'] = best_pattern

            for pid, card_list in trick['played']:
                for card in card_list:
                    if card.rank in SCORE_RANKS:
                        trick['score_cards'].append(card)
                        trick['score'] += SCORE_VALUES[card.rank]

            rec.tricks.append(trick)

            parts = []
            for pid, card_list in trick['played']:
                cs = cards_str(card_list)
                parts.append(f"玩{pid+1}:{cs}")
            ci = ' | '.join(parts)
            pattern_name = {'single': '单张', '510k': '5·10·K',
                            'hong': '轰', 'zha': '炸'}.get(trick['pattern'], '单张')
            rec.log(f"第{t}圈 [{pattern_name}]: 玩{leader+1}首出 → [{ci}] → 赢: 玩{best_pid+1}({cards_str(best_cards)}) 分={trick['score']}")
            leader = best_pid

        rec.attacker_score = sum(tr['score'] for tr in rec.tricks)
        if rec.tricks:
            lt = rec.tricks[-1]
            rec.last_trick_winner_pid = lt['winner']
            rec.last_trick_winner_side = lt['winner_side']
            for pid, card_list in lt['played']:
                if pid == lt['winner']:
                    rec.last_trick_card = card_list[-1] if card_list else None
                    break
        rec.log(f"抓分方总分={rec.attacker_score}（共{len(rec.tricks)}圈）")

    def _settle(self, rec):
        sc = rec.attacker_score
        is_bottom = rec.last_trick_winner_side == 'attacker'

        if sc == 0:
            rec.result_title = '光头'; rec.base_up_att = 0; rec.final_up_def = 3
        elif sc <= 35:
            rec.result_title = '干受苦'; rec.base_up_att = 0; rec.final_up_def = 1
        elif sc <= 45:
            rec.result_title = '上台'; rec.base_up_att = 0; rec.final_up_def = 0
        else:
            rec.base_up_att = min((sc - 50) // 10 + 1, 6)
            rec.final_up_def = 0
            rec.result_title = f"升{rec.base_up_att}级"

        bonus = 0
        if is_bottom:
            bonus = 4 if (rec.last_trick_card and rec.last_trick_card.rank == '大王') else 3
        rec.bonus_up = bonus

        if is_bottom and sc < 40:
            rec.result_title = '干扣底'
            rec.final_up_def = 0; rec.base_up_att = 0; rec.bonus_up = 0

        rec.final_up_att = rec.base_up_att + bonus

        old_def, old_att = self.defender_level, self.attacker_level
        old_ta, old_tb = self.team_a_level, self.team_b_level  # 记录过7前的等级
        self.team_a_level_before_over7 = old_ta
        self.team_b_level_before_over7 = old_tb
        self.defender_level = level_up(self.defender_level, rec.final_up_def)
        self.attacker_level = level_up(self.attacker_level, rec.final_up_att)

        # 按队伍追踪等级（队伍A=初始庄家方=玩家0+2）
        if self.dealer_pid in (0, 2):
            # 当前庄家是队伍A
            self.team_a_level = level_up(self.team_a_level, rec.final_up_def)
            self.team_b_level = level_up(self.team_b_level, rec.final_up_att)
        else:
            # 当前庄家是队伍B
            self.team_b_level = level_up(self.team_b_level, rec.final_up_def)
            self.team_a_level = level_up(self.team_a_level, rec.final_up_att)

        # 累计升级步数（跨局累计，用于过7判定）
        if self.dealer_pid in (0, 2):
            self.team_a_cumulative_steps += rec.final_up_def
            self.team_b_cumulative_steps += rec.final_up_att
        else:
            self.team_b_cumulative_steps += rec.final_up_def
            self.team_a_cumulative_steps += rec.final_up_att

        rec.log(f"结算: 抓分={sc} 扣底={'是' if is_bottom else '否'} 庄方+{rec.final_up_def}({old_def}→{self.defender_level}) 抓方+{rec.final_up_att}({old_att}→{self.attacker_level})")

        # 守庄局：双方级牌锁定在7，不随结算升级变动，只累计步数
        if self.defending_team:
            self.team_a_level = '7'
            self.team_b_level = '7'
            self.defender_level = '7'
            self.attacker_level = '7'

        self._check_over7(rec)

    def _check_over7(self, rec):
        rec.game_over_check = True

        LEVEL_CYCLE_LEN = 10

        def _has_over7(cumulative_steps, current_level):
            return cumulative_steps >= LEVEL_CYCLE_LEN and level_idx(current_level) > 0

        # === 守庄局优先处理 ===
        if self.defending_team:
            is_defender_A = self.defending_team == 'A'
            def_cum = self.team_a_cumulative_steps if is_defender_A else self.team_b_cumulative_steps
            def_lvl = self.team_a_level if is_defender_A else self.team_b_level
            opp_cum = self.team_b_cumulative_steps if is_defender_A else self.team_a_cumulative_steps
            opp_lvl = self.team_b_level if is_defender_A else self.team_a_level
            dname = f'队伍{self.defending_team}'
            oname = '队伍B' if is_defender_A else '队伍A'

            # 1) 守庄成功
            if rec.attacker_score <= 35:
                rec.result_title = f'{dname}守庄成功🏆'
                rec.round_ended = True
                rec.round_winner = dname
                if not self.total_rounds:
                    self.game_over = True
                    self.winner = f'{dname}（守庄方）'
                rec.log(f"🏆 {dname}守庄成功（对方得分{rec.attacker_score}≤35），本轮胜利！")
                if self.total_rounds:
                    self._reset_for_new_round()
                return
            # 2) 守庄方过7
            if _has_over7(def_cum, def_lvl):
                rec.result_title = f'{dname}过7🏆'
                rec.round_ended = True
                rec.round_winner = dname
                if not self.total_rounds:
                    self.game_over = True
                    self.winner = f'{dname}（守庄方）'
                rec.log(f"🏆 {dname}（守庄方）过7，本轮胜利！")
                if self.total_rounds:
                    self._reset_for_new_round()
                return
            # 3) 对方过7
            if _has_over7(opp_cum, opp_lvl):
                rec.result_title = f'{oname}过7🏆'
                rec.round_ended = True
                rec.round_winner = oname
                if not self.total_rounds:
                    self.game_over = True
                    self.winner = f'{oname}（庄家方）'
                rec.log(f"🏆 {oname}过7，守庄失败！")
                if self.total_rounds:
                    self._reset_for_new_round()
                return
            return

        # === 非守庄局：动态判断当前庄家 ===
        dealer_is_a = self.dealer_pid in (0, 2)
        dealer_cum = self.team_a_cumulative_steps if dealer_is_a else self.team_b_cumulative_steps
        dealer_lvl = self.team_a_level if dealer_is_a else self.team_b_level
        attacker_cum = self.team_b_cumulative_steps if dealer_is_a else self.team_a_cumulative_steps
        attacker_lvl = self.team_b_level if dealer_is_a else self.team_a_level
        dlabel = '队伍A' if dealer_is_a else '队伍B'
        alabel = '队伍B' if dealer_is_a else '队伍A'

        # 庄家方过7 → 直接获胜（单轮）/ 本轮结束（多轮）
        if _has_over7(dealer_cum, dealer_lvl):
            rec.result_title = f'{dlabel}过7🏆'
            rec.round_ended = True
            rec.round_winner = dlabel
            if not self.total_rounds:
                self.game_over = True
                self.winner = f'{dlabel}（庄家方）'
            rec.log(f"🏆 {dlabel}（庄家方）过7（累计{dealer_cum}步），本轮结束！")
            if self.total_rounds:
                self._reset_for_new_round()
            return

        # 抓分方过7 → 守庄
        if _has_over7(attacker_cum, attacker_lvl):
            if dealer_is_a:
                self.team_b_level_before_over7 = self.team_b_level
            else:
                self.team_a_level_before_over7 = self.team_a_level
            self.team_b_level = '7'
            self.team_a_level = '7'
            new_dealer = rec.attacker_team[0]
            if new_dealer != self.dealer_pid:
                self.dealer_pid = new_dealer
                self.defender_level = '7'
                self.attacker_level = '7'
            rec.result_title = f'{alabel}过7🏆'
            rec.round_ended = True
            rec.round_winner = alabel
            if not self.total_rounds:
                self.defending_team = 'B' if dealer_is_a else 'A'
                rec.result_title = f'{alabel}过7→守庄🏰'
                rec.round_winner = f'{alabel}（进入守庄）'
                rec.log(f"🏰 {alabel}（抓分方）过7，进入守庄！")
            else:
                self._reset_for_new_round()
            rec.log(f"🏆 {alabel}过7，本轮结束！")
            return

        # 庄权交换
        if rec.result_title == '上台' or rec.final_up_att > 0:
            new_dealer = rec.attacker_team[0]
            if new_dealer != self.dealer_pid:
                self.dealer_pid = new_dealer
                self.defender_level, self.attacker_level = self.attacker_level, self.defender_level
                rec.log(f"庄家变更: → 玩{self.dealer_pid+1} | 级牌: 庄方={self.defender_level} 抓方={self.attacker_level}")

    def _reset_for_new_round(self):
        self.team_a_cumulative_steps = 0
        self.team_b_cumulative_steps = 0
        self.defending_team = None
        self.team_a_level = '7'
        self.team_b_level = '7'
        self.defender_level = '7'
        self.attacker_level = '7'



# ==================== Excel ====================

def save_excel(records, game, path):
    wb = Workbook()
    tf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    tfont = Font(size=14, bold=True, color="FFFFFF")
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    sf = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sfont = Font(bold=True, size=11)

    # ====== Sheet1: 总览 ======
    ws = wb.active; ws.title = "游戏总览"
    ws.merge_cells('A1:P1')
    c = ws['A1']; c.value = f"一副牌升级游戏模拟（过7）"; c.font = tfont; c.fill = tf; c.alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:P2')
    total_games = len(records)
    total_r = len(game.round_records) if hasattr(game, 'round_records') else 0
    c = ws['A2']; c.value = f"生成: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 共{total_r}轮{total_games}局 | 胜方: {game.winner or '未完成'}"; c.font = Font(italic=True)

    hdrs = ['轮次','局数','庄家','庄方级(前)','抓方级(前)','本局级',
            '定主方式','亮牌/闷牌','主花色',
            '抓分','扣底','扣底牌',
            '抓方升级','庄方升级','庄方级(后)','抓方级(后)','结果']
    r = 4
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = hfont; c.fill = hf; c.alignment = Alignment(horizontal='center')

    cur_def, cur_att = '7', '7'
    for i, rec in enumerate(records):
        r = 5 + i
        pre_def, pre_att = cur_def, cur_att

        if rec.trump_method == 'bright':
            tm, td = '亮牌', f"玩{rec.bright_pid+1}亮{rec.bright_card}"
        elif rec.trump_method == 'concealed':
            tm, td = '闷牌', f"玩{rec.concealed_pid+1}闷{rec.concealed_card}"
        else:
            tm, td = '底牌首张', '底牌首张定主'
        ts = SUIT_CN.get(rec.trump_suit, rec.trump_suit or '—')
        ib = rec.last_trick_winner_side == 'attacker' if rec.last_trick_winner_side else False

        cur_def = level_up(cur_def, rec.final_up_def)
        cur_att = level_up(cur_att, rec.final_up_att)

        # 找本轮的 round 编号
        rnd_num = None
        if hasattr(game, 'round_records'):
            for rr in game.round_records:
                if rr['start_rnd'] <= rec.rnd <= rr['end_rnd']:
                    rnd_num = rr['round']
                    break
            if rnd_num is None:
                rnd_num = game.current_round - 1 if hasattr(game, 'current_round') else '—'

        vals = [f"第{rnd_num}轮" if rnd_num else '—', rec.rnd, f"玩{rec.dealer_pid+1}", pre_def, pre_att, rec.level,
                tm, td, ts, rec.attacker_score,
                '是' if ib else '否', str(rec.last_trick_card or '—'),
                f"+{rec.final_up_att}" if rec.final_up_att else 0,
                f"+{rec.final_up_def}" if rec.final_up_def else 0,
                cur_def, cur_att, rec.result_title]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.alignment = Alignment(horizontal='center')
        # 每轮最后一行高亮
        if hasattr(game, 'round_records'):
            for rr in game.round_records:
                if rec.rnd == rr['end_rnd']:
                    for col in range(1, len(hdrs)+1):
                        ws.cell(row=r, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(hdrs[col-1])*2)

    # ====== Sheet2: 轮次统计 ======
    ws2 = wb.create_sheet("轮次统计")
    ws2.merge_cells('A1:E1')
    c = ws2['A1']; c.value = "轮次统计"; c.font = tfont; c.fill = tf; c.alignment = Alignment(horizontal='center')
    r = 2
    hdrs2 = ['轮次', '起始局', '结束局', '局数', '本轮胜方']
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=r, column=col, value=h)
        c.font = hfont; c.fill = hf; c.alignment = Alignment(horizontal='center')
    r = 3
    for rr in game.round_records:
        vals = [f"第{rr['round']}轮", rr['start_rnd'], rr['end_rnd'], rr['games_count'], rr['winner']]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=r, column=col, value=v).alignment = Alignment(horizontal='center')
        r += 1
    # 汇总行
    if game.round_records:
        total_g = sum(rr['games_count'] for rr in game.round_records)
        for col, v in enumerate(['合计', '', '', total_g, game.winner or '—'], 1):
            c = ws2.cell(row=r, column=col, value=v)
            c.alignment = Alignment(horizontal='center')
            c.font = Font(bold=True)
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # ====== Sheet3: 每局详情 ======
    ws3 = wb.create_sheet("每局详情")
    row = 1

    for rec in records:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws3.cell(row=row, column=1, value=f"═══ 第{rec.rnd}局 ─ {rec.result_title} ═══")
        c.font = tfont; c.fill = tf; row += 1

        for label, val in [
            ("庄家", f"玩家{rec.dealer_pid+1}"),
            ("庄家方级牌", str(rec.defender_level)),
            ("抓分方级牌", str(rec.attacker_level)),
            ("本局级牌", str(rec.level)),
            ("庄家阵营", f"玩家{rec.dealer_team[0]+1}、玩家{rec.dealer_team[1]+1}"),
            ("抓分阵营", f"玩家{rec.attacker_team[0]+1}、玩家{rec.attacker_team[1]+1}"),
            ("定主方式", rec.trump_method or '—'),
            ("主花色", SUIT_CN.get(rec.trump_suit, rec.trump_suit or '—') if rec.trump_suit else '—'),
        ]:
            c = ws3.cell(row=row, column=1, value=label); c.font = sfont
            ws3.cell(row=row, column=3, value=val); row += 1

        if rec.bright_pid is not None:
            c = ws3.cell(row=row, column=1, value="⭐ 亮牌"); c.font = sfont; c.fill = sf; row += 1
            ws3.cell(row=row, column=1, value="亮牌玩家"); ws3.cell(row=row, column=2, value=f"玩家{rec.bright_pid+1}")
            ws3.cell(row=row, column=3, value="亮出的牌"); ws3.cell(row=row, column=4, value=str(rec.bright_card)); row += 1
        if rec.concealed_pid is not None:
            c = ws3.cell(row=row, column=1, value="🃏 闷牌"); c.font = sfont; c.fill = sf; row += 1
            ws3.cell(row=row, column=1, value="闷牌玩家"); ws3.cell(row=row, column=2, value=f"玩家{rec.concealed_pid+1}")
            ws3.cell(row=row, column=3, value="闷的牌"); ws3.cell(row=row, column=4, value=str(rec.concealed_card)); row += 1

        c = ws3.cell(row=row, column=1, value="📋 初始手牌"); c.font = sfont; c.fill = sf; row += 1
        for pid in range(4):
            ws3.cell(row=row, column=1, value=f"玩家{pid+1}")
            ws3.cell(row=row, column=2, value=cards_str(rec.initial_hands.get(pid, [])))
            ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8); row += 1

        c = ws3.cell(row=row, column=1, value="📦 底牌信息"); c.font = sfont; c.fill = sf; row += 1
        ws3.cell(row=row, column=1, value="初始底牌"); ws3.cell(row=row, column=2, value=cards_str(rec.initial_bottom))
        ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8); row += 1

        if rec.buried_cards:
            ws3.cell(row=row, column=1, value="庄家埋入"); ws3.cell(row=row, column=2, value=cards_str(rec.buried_cards))
            ws3.cell(row=row, column=3, value="埋底后底牌"); ws3.cell(row=row, column=4, value=cards_str(rec.bottom_after_bury))
            bs = sum(SCORE_VALUES.get(c.rank,0) for c in rec.bottom_after_bury)
            ws3.cell(row=row, column=5, value=f"分值={bs}")
            ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3); row += 1

        if rec.picked_from_bottom:
            ws3.cell(row=row, column=1, value="拣出主牌"); ws3.cell(row=row, column=2, value=cards_str(rec.picked_from_bottom))
            ws3.cell(row=row, column=3, value="弃回底牌"); ws3.cell(row=row, column=4, value=cards_str(rec.discarded_to_bottom))
            ws3.cell(row=row, column=5, value="捡主后底牌"); ws3.cell(row=row, column=6, value=cards_str(rec.bottom_after_pick)); row += 1

        c = ws3.cell(row=row, column=1, value="📝 操作日志"); c.font = sfont; c.fill = sf; row += 1
        for log in rec.logs:
            ws3.cell(row=row, column=1, value=log)
            ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8); row += 1

        c = ws3.cell(row=row, column=1, value="📊 结算"); c.font = sfont; c.fill = sf; row += 1
        ib = rec.last_trick_winner_side == 'attacker' if rec.last_trick_winner_side else False
        for label, val in [
            ("抓分方得分", str(rec.attacker_score)),
            ("是否扣底", "是" if ib else "否"),
            ("扣底牌", str(rec.last_trick_card) if rec.last_trick_card else "—"),
            ("基础升级 抓方", f"+{rec.base_up_att}" if rec.base_up_att else "0"),
            ("扣底加成", f"+{rec.bonus_up}" if rec.bonus_up else "无"),
            ("总升级 抓方", f"+{rec.final_up_att}"),
            ("总升级 庄方", f"+{rec.final_up_def}"),
            ("结果", rec.result_title),
        ]:
            ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=row, column=3, value=val); row += 1

        row += 2

    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path)
    print(f"✅ Excel: {path}")

# ==================== 主程序 ====================

def main():
    import argparse
    p = argparse.ArgumentParser(description='一副牌升级游戏模拟器')
    p.add_argument('--seed', type=int, default=None, help='随机种子（可复现）')
    p.add_argument('--rounds', type=int, default=None, help='总轮数（默认不限制，打到200局上限）')
    p.add_argument('--max-games', type=int, default=None, help='总局数上限（默认200，设0无限制）')
    args = p.parse_args()

    print("🎮 一副牌升级游戏模拟器")
    print("=" * 50)
    if args.rounds:
        print(f"📊 目标轮数: {args.rounds} 轮")

    max_games = 0 if args.rounds else None  # 指定轮数时不设局数上限
    game = Game(total_rounds=args.rounds, max_games=max_games)
    if args.seed:
        import random
        random.seed(args.seed)
    records = game.run()

    print(f"\n游戏结束! 共{len(game.round_records)}轮{len(records)}局 | 胜方: {game.winner or '未完成'}")
    print(f"最终: 庄家方={game.defender_level} 抓分方={game.attacker_level}")

    # 打印轮次统计
    print("\n📊 轮次统计:")
    for rr in game.round_records:
        print(f"  第{rr['round']}轮: 局{rr['start_rnd']}~{rr['end_rnd']} 共{rr['games_count']}局 | 胜方: {rr['winner']}")

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, f'一副牌升级游戏模拟_{ts}.xlsx')
    save_excel(records, game, path)

    print(f"\n🏆 最终胜方: {game.winner or '未完成'}")

if __name__ == '__main__':
    main()
