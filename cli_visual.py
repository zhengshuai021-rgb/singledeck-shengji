#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
一副牌升级 - CLI 可视化模拟器
输入基础参数 → 逐步观看出牌 → 最终结算

创建时间：2026-05-28
作者：Kami 🐱
"""

import random
import sys
import time
import os
import re
import argparse
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ==================== 颜色 & 样式 ====================

class C:
    """ANSI 颜色代码"""
    R = '\033[0m'       # reset
    BOLD = '\033[1m'
    DIM = '\033[2m'
    RED = '\033[91m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    WHITE = '\033[97m'
    BG_BLUE = '\033[44m'
    BG_GREEN = '\033[42m'
    BG_RED = '\033[41m'
    BG_YELLOW = '\033[43m'
    BG_MAGENTA = '\033[45m'
    BG_DARK = '\033[48;5;235m'

    # 花色颜色
    @staticmethod
    def suit_color(suit):
        if suit in ('♥', '♦'):
            return C.RED
        return C.WHITE

    @staticmethod
    def team_color(pid, dealer_pid):
        """庄家方蓝色，抓分方绿色"""
        team = 'dealer' if pid in (dealer_pid, (dealer_pid + 2) % 4) else 'attacker'
        return C.CYAN if team == 'dealer' else C.GREEN

    @staticmethod
    def player_label(pid, dealer_pid):
        tc = C.team_color(pid, dealer_pid)
        team = '庄' if pid in (dealer_pid, (dealer_pid + 2) % 4) else '抓'
        return f"{tc}玩{pid+1}[{team}]{C.R}"

def card_str(card, highlight=False):
    """单张牌，带花色颜色"""
    if card.rank in ('大王', '小王'):
        prefix = C.YELLOW if card.rank == '大王' else C.MAGENTA
        h = C.BG_YELLOW if highlight else ''
        return f"{h}{prefix}{card.rank}{C.R}"
    sc = C.suit_color(card.suit)
    h = C.BG_YELLOW if highlight else ''
    return f"{h}{sc}{card.suit}{card.rank}{C.R}"

def cards_str(cards, highlight_set=None):
    """多张牌"""
    if not cards:
        return f"{C.DIM}无{C.R}"
    if highlight_set is None:
        highlight_set = set()
    parts = []
    for c in cards:
        hl = c in highlight_set
        parts.append(card_str(c, highlight=hl))
    return ' '.join(parts)

def _strip_ansi(s):
    """去除 ANSI 颜色码"""
    return re.sub(r'\x1b\[[0-9;]*m', '', s)

def cards_str_plain(cards):
    """纯文本牌串（用于 Excel 输出）"""
    if not cards:
        return '无'
    return ' '.join(str(c) for c in cards)

def delay(ms):
    """等待（毫秒），--fast 时跳过"""
    if not _FAST:
        time.sleep(ms / 1000.0)

def clear():
    os.system('clear' if os.name != 'nt' else 'cls')

def separator(char='═', width=70):
    print(f"\n{C.DIM}{char * width}{C.R}")

def box(title, lines, border_color=C.CYAN):
    """打印一个框"""
    w = 70
    print(f"{border_color}╔{'═' * (w-2)}╗{C.R}")
    print(f"{border_color}║{C.R} {C.BOLD}{title}{C.R}")
    print(f"{border_color}╠{'═' * (w-2)}╣{C.R}")
    for line in lines:
        print(f"{border_color}║{C.R} {line}")
    print(f"{border_color}╚{'═' * (w-2)}╝{C.R}")

def wait_prompt(text="按回车继续..."):
    """等待用户按键"""
    if not _FAST:
        try:
            input(f"\n  {C.DIM}{text}{C.R}")
        except (EOFError, KeyboardInterrupt):
            pass
    else:
        # fast 模式：打印分隔
        print(f"  {C.DIM}---{C.R}")

# ==================== 全局配置 ====================

_FAST = False

# ==================== 常量 ====================

SUITS = ['♠', '♥', '♣', '♦']
RANKS = ['2', '3', '4', '5', '6', '7', '8', '9', '10', 'J', 'Q', 'K', 'A']
SCORE_RANKS = {'5', '10', 'K'}
SCORE_VALUES = {'5': 5, '10': 10, 'K': 10}
RANK_ORDER = {rank: idx for idx, rank in enumerate(RANKS)}
SUIT_CN = {'♠': '黑桃', '♥': '红桃', '♣': '草花', '♦': '方块'}
PATTERN_NAMES = {'single': '单张', '510k': '5·10·K', 'hong': '💥 轰', 'zha': '💣 炸'}

LEVEL_CYCLE = ['7', '8', '9', '10', 'J', 'Q', 'K', 'A', '2', '3', '4', '5', '6']
LEVEL_CYCLE_LEN = len(LEVEL_CYCLE)

def level_idx(lvl):
    if lvl in LEVEL_CYCLE:
        return LEVEL_CYCLE.index(lvl)
    return 0

def level_up(lvl, steps=1):
    idx = level_idx(lvl)
    return LEVEL_CYCLE[(idx + steps) % LEVEL_CYCLE_LEN]

# ==================== 牌类 ====================

class Card:
    def __init__(self, suit, rank):
        self.suit = suit
        self.rank = rank
    def __repr__(self):
        if self.rank in ('大王', '小王'): return self.rank
        return f"{self.suit}{self.rank}"
    def __eq__(self, other):
        return isinstance(other, Card) and self.suit == other.suit and self.rank == other.rank
    def __hash__(self):
        return hash((self.suit, self.rank))

def create_deck():
    deck = []
    for suit in SUITS:
        for rank in RANKS:
            deck.append(Card(suit, rank))
    deck.append(Card('王', '大王'))
    deck.append(Card('王', '小王'))
    return deck

# ==================== 牌力判定 ====================

def is_main(card, level, trump_suit):
    """判断是否主牌（常主 + 主花色牌）"""
    if card.rank in ('大王', '小王'): return True
    if card.rank == '3' and card.suit == '♥': return True
    if card.rank == '2': return True
    if card.rank == level: return True
    if card.suit == trump_suit: return True
    return False

def card_power(card, level, trump_suit):
    if card.rank == '大王': return (5, 100)
    if card.rank == '小王': return (4, 100)
    if card.rank == level: return (3, RANK_ORDER[card.rank])
    if card.rank == '3' and card.suit == '♥': return (2, 100)
    if card.rank == '2': return (2, 90)
    if card.suit == trump_suit: return (1, RANK_ORDER[card.rank])
    return (0, RANK_ORDER[card.rank])

def cp(card, level, trump_suit):
    g, v = card_power(card, level, trump_suit)
    return g * 1000 + v

def compare_cards(c1, c2, level, trump_suit, lead_suit):
    c1_lead = c1.suit == lead_suit
    c2_lead = c2.suit == lead_suit
    c1_main = is_main(c1, level, trump_suit)
    c2_main = is_main(c2, level, trump_suit)
    if c1_lead and c2_lead:
        p1, p2 = card_power(c1, level, trump_suit), card_power(c2, level, trump_suit)
        if p1[0] != p2[0]: return 1 if p1[0] > p2[0] else -1
        return 1 if p1[1] > p2[1] else (-1 if p1[1] < p2[1] else 1)
    if c1_lead and not c2_lead:
        return -1 if c2_main else 1
    if not c1_lead and c2_lead:
        return 1 if c1_main else -1
    if c1_main and c2_main:
        p1, p2 = card_power(c1, level, trump_suit), card_power(c2, level, trump_suit)
        if p1[0] != p2[0]: return 1 if p1[0] > p2[0] else -1
        return 1 if p1[1] > p2[1] else (-1 if p1[1] < p2[1] else 1)
    if c1_main: return 1
    if c2_main: return -1
    return 1

def group_by_rank(cards):
    d = defaultdict(list)
    for c in cards: d[c.rank].append(c)
    return d

def find_510k(hand, suit):
    suit_cards = [c for c in hand if c.suit == suit and c.rank in ('5', '10', 'K')]
    ranks_present = {c.rank for c in suit_cards}
    if ranks_present == {'5', '10', 'K'}:
        return [c for c in suit_cards]
    return None

def find_hongs(hand):
    groups = group_by_rank(hand)
    return [(rank, cards) for rank, cards in groups.items() if len(cards) == 4]

def find_zhas(hand):
    groups = group_by_rank(hand)
    sa_list = [c for c in groups.get('A', []) if c.suit == '♠']
    if not sa_list:
        return []
    result = []
    for rank, cards in groups.items():
        if rank == 'A':
            continue
        if len(cards) >= 3:
            result.append((rank, [sa_list[0]] + cards[:3]))
    return result

def count_hand_patterns(hand):
    return {
        'hong': len(find_hongs(hand)),
        'zha': len(find_zhas(hand)),
        '510k': sum(1 for s in SUITS if find_510k(hand, s)),
    }

def check_deal_requirements(hands, requirements):
    for pattern, (lo, hi) in requirements.items():
        side_a_ok = any(lo <= count_hand_patterns(h)[pattern] <= hi for h in (hands[0], hands[2]))
        side_b_ok = any(lo <= count_hand_patterns(h)[pattern] <= hi for h in (hands[1], hands[3]))
        if not side_a_ok or not side_b_ok:
            return False
    return True

def pattern_hierarchy(pattern, cards, level, trump_suit):
    if pattern == '510k':
        is_main_510k = all(is_main(c, level, trump_suit) for c in cards)
        return (4 if is_main_510k else 3, 0)
    if pattern == 'hong':
        return (2, RANK_ORDER.get(cards[0].rank, 0))
    if pattern == 'zha':
        r = cards[1].rank if cards[0].rank == 'A' else cards[0].rank
        return (1, RANK_ORDER.get(r, 0))
    return (0, 0)

def compare_trick_patterns(pattern_a, cards_a, pattern_b, cards_b, level, trump_suit, lead_suit):
    ha = pattern_hierarchy(pattern_a, cards_a, level, trump_suit)
    hb = pattern_hierarchy(pattern_b, cards_b, level, trump_suit)
    if ha[0] != hb[0]:
        return 1 if ha[0] > hb[0] else -1
    if ha[0] == 0:
        return compare_cards(cards_a[0], cards_b[0], level, trump_suit, lead_suit)
    if ha[0] in (4, 3):
        max_a = max(RANK_ORDER.get(c.rank, 0) for c in cards_a)
        max_b = max(RANK_ORDER.get(c.rank, 0) for c in cards_b)
        return 1 if max_a > max_b else (-1 if max_a < max_b else 1)
    if ha[0] == 2:
        return 1 if RANK_ORDER.get(cards_a[0].rank, 0) > RANK_ORDER.get(cards_b[0].rank, 0) else -1
    if ha[0] == 1:
        r_a = cards_a[1].rank if cards_a[0].rank == 'A' else cards_a[0].rank
        r_b = cards_b[1].rank if cards_b[0].rank == 'A' else cards_b[0].rank
        return 1 if RANK_ORDER.get(r_a, 0) > RANK_ORDER.get(r_b, 0) else -1
    return 1

def max_card_in_trick(played, level, trump_suit, lead_suit):
    flat = []
    for item in played:
        pid = item[0]
        cards_or_card = item[1]
        if isinstance(cards_or_card, list):
            for card in cards_or_card:
                flat.append((pid, card))
        else:
            flat.append((pid, cards_or_card))
    if not flat:
        return None, None
    best_pid, best_card = flat[0]
    for pid, card in flat[1:]:
        if compare_cards(card, best_card, level, trump_suit, lead_suit) == 1:
            best_pid, best_card = pid, card
    return best_pid, best_card

# ==================== 机器人 ====================

class Bot:
    def __init__(self, pid, hand, side, level, trump_suit):
        self.pid = pid
        self.hand = list(hand)
        self.side = side
        self.level = level
        self.trump_suit = trump_suit

    def _non_score_offsuit(self):
        return [c for c in self.hand if not is_main(c, self.level, self.trump_suit)
                and c.rank not in SCORE_RANKS]

    def _all_main(self):
        return [c for c in self.hand if is_main(c, self.level, self.trump_suit)]

    def _detect_pattern(self, cards):
        if len(cards) == 1:
            return 'single'
        if len(cards) == 3:
            if {c.rank for c in cards} == {'5', '10', 'K'}:
                return '510k'
        if len(cards) == 4:
            ranks = [c.rank for c in cards]
            if len(set(ranks)) == 1:
                return 'hong'
            if len(set(ranks)) == 2 and 'A' in ranks and ranks.count('A') == 1:
                return 'zha'
        return 'single'

    def lead(self):
        """首出：优先轰 > 510K > 单张"""
        level, ts = self.level, self.trump_suit
        if not self.hand:
            return []

        # 1. 优先出轰（副牌轰优先，无副牌轰则出主牌轰）
        off_hong = None
        main_hong = None
        for rank, cards in find_hongs(self.hand):
            if not all(is_main(c, level, ts) for c in cards):
                if off_hong is None:
                    off_hong = cards
            else:
                if main_hong is None:
                    main_hong = cards
        if off_hong:
            for c in off_hong: self.hand.remove(c)
            return off_hong
        if main_hong:
            for c in main_hong: self.hand.remove(c)
            return main_hong

        # 2. 出510K（副牌优先，无副牌则出主牌510K）
        off_510k = None
        main_510k = None
        for suit in SUITS:
            cards = find_510k(self.hand, suit)
            if cards:
                if suit == ts:
                    if main_510k is None:
                        main_510k = cards
                else:
                    if off_510k is None:
                        off_510k = cards
        if off_510k:
            for c in off_510k: self.hand.remove(c)
            return off_510k
        if main_510k:
            for c in main_510k: self.hand.remove(c)
            return main_510k

        # 3. 单张
        safe = self._non_score_offsuit()
        if safe:
            safe.sort(key=lambda c: RANK_ORDER.get(c.rank, 0))
            card = safe[0]
        else:
            ns = [c for c in self.hand if c.rank not in SCORE_RANKS]
            card = min(ns if ns else self.hand, key=lambda c: RANK_ORDER.get(c.rank, 0))
        self.hand.remove(card)
        return [card]

    def follow(self, lead_suit, played_so_far, is_last_trick=False):
        level, ts = self.level, self.trump_suit
        if not self.hand:
            return []
        first_cards = played_so_far[0][1] if played_so_far else []
        if not first_cards:
            return self._discard_or_trump(played_so_far, is_last_trick, need=1)
        need_count = len(first_cards)
        if len(self.hand) <= need_count:
            out = list(self.hand)
            self.hand.clear()
            return out
        lead_pattern = self._detect_pattern(first_cards)
        if lead_pattern == '510k':
            all_main = self._all_main()
            main_510k = find_510k(self.hand, ts)
            supreme = [c for c in self.hand if c.rank in ('大王', '小王')
                       or (c.rank == '3' and c.suit == '♥')]
            if len(supreme) >= 3 and main_510k is None:
                out = supreme[:3]
                for c in out: self.hand.remove(c)
                return out
            if main_510k:
                for c in main_510k: self.hand.remove(c)
                return main_510k
            if all_main:
                all_main.sort(key=lambda c: cp(c, level, ts))
                out = all_main[:need_count]
                for c in out: self.hand.remove(c)
                if len(out) < need_count:
                    off = [c for c in self.hand if not is_main(c, level, ts)]
                    padding = off[:need_count - len(out)]
                    out += padding
                    for c in padding: self.hand.remove(c)
                return out
            off = [c for c in self.hand if not is_main(c, level, ts)]
            out = off[:need_count]
            if len(out) < need_count:
                remaining = [c for c in self.hand if c not in out]
                out += remaining[:need_count - len(out)]
            for c in out: self.hand.remove(c)
            return out
        if lead_pattern == 'hong':
            lead_rank = first_cards[0].rank
            lead_hong_main = all(is_main(c, level, ts) for c in first_cards)
            for rank, cards in find_hongs(self.hand):
                card_main = all(is_main(c, level, ts) for c in cards)
                can_beat = False
                if card_main and not lead_hong_main:
                    can_beat = True
                elif card_main == lead_hong_main:
                    can_beat = RANK_ORDER.get(rank, 0) > RANK_ORDER.get(lead_rank, 0)
                if can_beat:
                    for c in cards: self.hand.remove(c)
                    return cards
            zhas = find_zhas(self.hand)
            if zhas:
                rank, cards = zhas[0]
                for c in cards: self.hand.remove(c)
                return cards
            all_main = self._all_main()
            if all_main:
                all_main.sort(key=lambda c: cp(c, level, ts))
                out = all_main[:need_count]
                for c in out: self.hand.remove(c)
                if len(out) < need_count:
                    others = list(self.hand[:need_count - len(out)])
                    out += others
                    for c in others: self.hand.remove(c)
                return out
            out = list(self.hand[:need_count])
            for c in out: self.hand.remove(c)
            return out
        if lead_pattern == 'zha':
            for rank, cards in find_hongs(self.hand):
                for c in cards: self.hand.remove(c)
                return cards
            return self._discard_or_trump(played_so_far, is_last_trick, need=len(first_cards))
        same = [c for c in self.hand if c.suit == lead_suit]
        has_score = any(c.rank in SCORE_RANKS
                        for _, cl in played_so_far for c in cl)
        if same:
            if has_score:
                best_card = self._find_best_to_win(same, played_so_far, level, ts, lead_suit)
                if best_card:
                    same.remove(best_card)
                    out = [best_card] + same[:need_count - 1]
                else:
                    pool = same
                    out = pool[:need_count]
            else:
                ns = [c for c in same if c.rank not in SCORE_RANKS]
                pool = ns if ns else same
                out = pool[:need_count]
            for c in out: self.hand.remove(c)
            if len(out) < need_count:
                extra = [c for c in self.hand if c not in out][:need_count - len(out)]
                out += extra
                for c in extra: self.hand.remove(c)
            return out
        else:
            return self._discard_or_trump(played_so_far, is_last_trick, need=need_count)

    def _discard_or_trump(self, played_so_far, is_last_trick, need=1):
        level, ts = self.level, self.trump_suit
        lead_suit = None
        for _, cl in played_so_far:
            if cl:
                lead_suit = cl[0].suit
                break
        if lead_suit is None:
            lead_suit = '♠'
        has_score = any(c.rank in SCORE_RANKS
                        for _, cl in played_so_far for c in cl)
        all_main = self._all_main()
        out = None
        if has_score:
            best_pid, best_card = max_card_in_trick(played_so_far, level, ts, lead_suit)
            if best_card:
                best_p = cp(best_card, level, ts)
                if self.side == 'dealer':
                    can_win = [c for c in all_main if cp(c, level, ts) > best_p]
                    if can_win:
                        can_win.sort(key=lambda c: cp(c, level, ts))
                        out = can_win[:need]
                    if out is None:
                        main = [c for c in all_main
                                if c.rank not in ('大王', '小王')
                                and not (c.rank == '3' and c.suit == '♥')]
                        if main:
                            main.sort(key=lambda c: RANK_ORDER.get(c.rank, 0))
                            out = main[:need]
                else:
                    can_win = [c for c in all_main if cp(c, level, ts) > best_p]
                    if can_win:
                        can_win.sort(key=lambda c: cp(c, level, ts))
                        out = can_win[:need]
        if out is None:
            off = [c for c in self.hand if not is_main(c, level, ts)]
            off.sort(key=lambda c: RANK_ORDER.get(c.rank, 0))
            out = off[:need]
            if not out:
                out = list(self.hand[:need])
        for c in out:
            if c in self.hand:
                self.hand.remove(c)
        if len(out) < need:
            extra = list(self.hand[:need - len(out)])
            for c in extra:
                self.hand.remove(c)
            out += extra
        return out

    def _find_best_to_win(self, same_suit_cards, played_so_far, level, ts, lead_suit):
        if not played_so_far: return None
        best_pid, best_card = max_card_in_trick(played_so_far, level, ts, lead_suit)
        if best_card is None: return None
        best_p = cp(best_card, level, ts)
        can_win = [c for c in same_suit_cards if cp(c, level, ts) > best_p]
        if can_win:
            can_win.sort(key=lambda c: cp(c, level, ts))
            return can_win[0]
        return None

    def select_for_bottom(self, count):
        def pri(c):
            return (int(is_main(c, self.level, self.trump_suit)),
                    int(c.rank in SCORE_RANKS),
                    RANK_ORDER.get(c.rank, 0))
        s = sorted(self.hand, key=pri)
        sel = s[:count]
        for c in sel: self.hand.remove(c)
        return sel

# ==================== CLI 可视化引擎 ====================

class CLIGame:
    def __init__(self, seed=None, start_level='7', max_rounds=200, total_rounds=None, max_games=None, deal_requirements=None):
        if seed is not None:
            random.seed(seed)
        self.defender_level = start_level
        self.attacker_level = start_level
        self.dealer_pid = random.randint(0, 3)
        self.deal_requirements = deal_requirements or {}
        self.rnd = 0
        self.max_rounds = max_rounds
        self.max_games = max_games if max_games is not None else max_rounds  # 总局数上限
        self.game_over = False
        self.winner = None
        self.team_a_level = start_level
        self.team_b_level = start_level
        self.team_a_cumulative_steps = 0
        self.team_b_cumulative_steps = 0
        self.defending_team = None  # 守庄方（None=None / 'A' / 'B'）
        # 轮次追踪
        self.total_rounds = total_rounds  # 总轮数（None=不限制）
        self.current_round = 1  # 当前第几轮
        self.round_starts_at = 1  # 本轮从第几局开始
        self.round_records = []  # [{round, start_rnd, end_rnd, winner, games_count}, ...]
        self.records = []  # 所有局记录

    def run(self):
        self._show_header()
        while not self.game_over and (self.max_games <= 0 or self.rnd < self.max_games):
            self.rnd += 1
            rec = self._play_round()
            self.records.append(rec)
            # 检查轮次是否结束
            if rec.round_ended:
                self.round_records.append({
                    'round': self.current_round,
                    'start_rnd': self.round_starts_at,
                    'end_rnd': self.rnd,
                    'winner': rec.round_winner or '—',
                    'games_count': self.rnd - self.round_starts_at + 1,
                })
                self.current_round += 1
                self.round_starts_at = self.rnd + 1
            # 检查是否达到总轮数
            if self.total_rounds and self.current_round > self.total_rounds:
                self.game_over = True
            if not self.game_over and self.rnd < self.max_rounds:
                wait_prompt(f"第 {self.rnd+1} 局开始")
        self._show_final_result()

    def _show_header(self):
        clear()
        lines = [
            f"  起始级牌: {C.BOLD}{self.defender_level}{C.R}",
            f"  最大局数: {self.max_rounds}",
            f"  模式: {'⚡ 快速' if _FAST else '🐢 逐步（按回车推进）'}",
        ]
        if self.deal_requirements:
            parts = [f"{k}({lo}-{hi})" for k, (lo, hi) in self.deal_requirements.items()]
            lines.append(f"  🃏 牌型要求: {', '.join(parts)}")
        box("🎮 一副牌升级 · CLI 可视化模拟器", lines)
        print()
        if not _FAST:
            wait_prompt("按回车开始游戏")
            clear()

    def _show_round_header(self, rec):
        separator()
        dt = rec.dealer_team
        at = rec.attacker_team

        # 队伍信息
        team_a = f"🔵 队伍A（玩1+玩3）: 级牌={C.BOLD}{self.team_a_level}{C.R}"
        team_b = f"🟢 队伍B（玩2+玩4）: 级牌={C.BOLD}{self.team_b_level}{C.R}"

        box(f"🃏 第 {self.rnd} 局", [
            f"  庄家: {C.player_label(self.dealer_pid, self.dealer_pid)}",
            f"  庄家阵营: {C.player_label(dt[0], self.dealer_pid)} {C.player_label(dt[1], self.dealer_pid)}",
            f"  抓分阵营: {C.player_label(at[0], self.dealer_pid)} {C.player_label(at[1], self.dealer_pid)}",
            f"  庄家方级牌: {C.BOLD}{self.defender_level}{C.R}  |  抓分方级牌: {C.BOLD}{self.attacker_level}{C.R}",
            f"  {team_a}",
            f"  {team_b}",
        ])
        print()

    def _show_hands(self, hands, bottom, rec):
        lines = []
        for pid in range(4):
            sorted_hand = sorted(hands[pid], key=lambda c: cp(c, rec.level, rec.trump_suit or ''), reverse=True)
            label = C.player_label(pid, self.dealer_pid)
            hand_str = cards_str(sorted_hand)
            lines.append(f"  {label} ({len(sorted_hand)}张): {hand_str}")
        lines.append(f"  {C.DIM}底牌 ({len(bottom)}张): {cards_str(bottom)}{C.R}")
        box("📋 初始手牌", lines)
        print()
        delay(300)
        if not _FAST:
            wait_prompt("查看手牌后，按回车发牌")

    def _show_trump(self, rec):
        dt = rec.dealer_team
        at = rec.attacker_team
        lvl = rec.level

        if rec.trump_method == 'bright':
            pid = rec.bright_pid
            card = rec.bright_card
            team = '庄家方' if pid in dt else '抓分方'
            box("⭐ 亮牌定主", [
                f"  {C.player_label(pid, self.dealer_pid)} 亮出了 {C.BOLD}{card_str(card, highlight=True)}{C.R}",
                f"  → 主花色: {C.BOLD}{C.suit_color(card.suit)}{SUIT_CN[card.suit]}{C.R}",
                f"  定主方: {team}",
            ])
        elif rec.trump_method == 'concealed':
            pid = rec.concealed_pid
            team = '庄家方' if pid in dt else '抓分方'
            box("🃏 闷牌定主", [
                f"  {C.player_label(pid, self.dealer_pid)} 暗扣了一张级牌",
                f"  → 主花色: {C.BOLD}?? (待揭晓){C.R}",
                f"  闷牌方: {team}",
            ])
        else:
            fc = next((c for c in rec.initial_bottom if c.suit in SUITS), rec.initial_bottom[0])
            ts = rec.trump_suit
            box("🎲 底牌首张定主", [
                f"  底牌首张: {card_str(fc, highlight=True)}",
                f"  → 主花色: {C.BOLD}{C.suit_color(ts)}{SUIT_CN.get(ts, ts)}{C.R}",
            ])
        print()
        delay(400)
        if not _FAST:
            wait_prompt("按回车继续 → 埋底")

    def _show_bury(self, rec):
        pid = self.dealer_pid
        lines = [
            f"  庄家 {C.player_label(pid, self.dealer_pid)} 选择埋底",
            f"  埋入: {cards_str(rec.buried_cards)}",
            f"  取回: {cards_str(list(set(rec.initial_bottom) - set(rec.bottom_after_bury)))}",
            f"  埋底后底牌: {cards_str(rec.bottom_after_bury)}",
        ]
        bs = sum(SCORE_VALUES.get(c.rank, 0) for c in rec.bottom_after_bury)
        score_color = C.RED if bs > 25 else C.GREEN
        lines.append(f"  底牌分值: {score_color}{bs}分{C.R}")
        box("📦 埋底", lines)
        print()
        delay(300)
        if not _FAST:
            wait_prompt("按回车继续")

    def _show_pick_main(self, rec):
        pid = rec.concealed_pid
        ts = rec.trump_suit
        lines = [
            f"  {C.player_label(pid, self.dealer_pid)} 翻开闷牌: {C.BOLD}{card_str(rec.concealed_card, highlight=True)}{C.R}",
            f"  → 主花色揭晓: {C.BOLD}{C.suit_color(ts)}{SUIT_CN.get(ts, ts)}{C.R}",
        ]
        if rec.picked_from_bottom:
            lines.append(f"  从底牌拣出: {cards_str(rec.picked_from_bottom)}")
            lines.append(f"  弃回底牌: {cards_str(rec.discarded_to_bottom)}")
            lines.append(f"  捡主后底牌: {cards_str(rec.bottom_after_pick)}")
        else:
            lines.append(f"  底牌无主牌，跳过")
        box("🔍 捡主", lines)
        print()
        delay(400)
        if not _FAST:
            wait_prompt("按回车开始出牌")

    def _show_trick(self, trick, rec, bots):
        dt = rec.dealer_team
        level = rec.level
        ts = rec.trump_suit

        pname = PATTERN_NAMES.get(trick.get('pattern', 'single'), '单张')

        separator('─', 60)
        print(f"  {C.BOLD}第 {trick['num']} 圈{C.R}  首出: {C.player_label(trick['leader'], self.dealer_pid)}  牌型: {pname}")
        print()

        # 显示每个玩家出的牌
        for pid, card_list in trick['played']:
            label = C.player_label(pid, self.dealer_pid)
            remaining = len(bots[pid].hand) if pid in bots else 0
            out_str = cards_str(card_list) if card_list else f"{C.DIM}（无牌）{C.R}"
            print(f"    {label} → {out_str}  {C.DIM}(剩余{remaining}张){C.R}")
            delay(150)

        # 赢家
        winner = trick.get('winner')
        if winner is not None:
            winner_cards = None
            for pid, cl in trick['played']:
                if pid == winner:
                    winner_cards = cl
                    break
            w_label = C.player_label(winner, self.dealer_pid)
            wc = cards_str(winner_cards) if winner_cards else ''
            score = trick.get('score', 0)
            score_str = f"{C.YELLOW}+{score}分{C.R}" if score > 0 else f"{C.DIM}0分{C.R}"
            print(f"\n    🏆 {w_label} 赢下本圈 {wc}  ({score_str})")
        print()
        delay(400)

    def _show_settle(self, rec):
        separator()
        sc = rec.attacker_score
        is_bottom = rec.last_trick_winner_side == 'attacker'

        # 用 round record 的旧值 + 升级量计算，不受 _check_over7 交换影响
        old_def = rec.defender_level
        old_att = rec.attacker_level
        new_def = level_up(old_def, rec.final_up_def)
        new_att = level_up(old_att, rec.final_up_att)

        lines = [
            f"  抓分方得分: {C.BOLD}{C.YELLOW if sc > 0 else C.RED}{sc}分{C.R}",
            f"  扣底: {'是' if is_bottom else '否'}",
            f"  结果: {C.BOLD}{rec.result_title}{C.R}",
        ]

        if rec.final_up_def > 0:
            lines.append(f"  {C.CYAN}庄家方{C.R} 升级: +{rec.final_up_def} ({old_def}→{new_def})")
        else:
            lines.append(f"  {C.CYAN}庄家方{C.R} 级牌: {new_def}（未升级）")

        if rec.final_up_att > 0:
            lines.append(f"  {C.GREEN}抓分方{C.R} 升级: +{rec.final_up_att} ({old_att}→{new_att})")
        else:
            lines.append(f"  {C.GREEN}抓分方{C.R} 级牌: {new_att}（未升级）")

        # 队伍等级（按当前庄/抓角色标注）
        cur_dealer_is_a = self.dealer_pid in (0, 2)
        lines.append("")
        if cur_dealer_is_a:
            lines.append(f"  🔵 {C.CYAN}队伍A（庄）{C.R} 级牌: {C.BOLD}{self.team_a_level}{C.R}")
            lines.append(f"  🟢 {C.GREEN}队伍B（抓）{C.R} 级牌: {C.BOLD}{self.team_b_level}{C.R}")
        else:
            lines.append(f"  🟢 {C.GREEN}队伍B（庄）{C.R} 级牌: {C.BOLD}{self.team_b_level}{C.R}")
            lines.append(f"  🔵 {C.CYAN}队伍A（抓）{C.R} 级牌: {C.BOLD}{self.team_a_level}{C.R}")

        box(f"📊 第 {self.rnd} 局结算", lines)
        print()
        delay(500)
        if not _FAST:
            wait_prompt("按回车继续")

    def _show_final_result(self):
        separator('═', 70)
        if self.winner:
            box("🏆 游戏结束", [
                f"  {C.BOLD}{C.YELLOW}胜方: {self.winner}{C.R}",
                f"  共进行了 {C.BOLD}{self.rnd}{C.R} 局",
                f"  🔵 队伍A 最终级牌: {self.team_a_level}",
                f"  🟢 队伍B 最终级牌: {self.team_b_level}",
            ])
        else:
            box("⏱️ 达到局数上限", [
                f"  进行了 {self.rnd} 局",
                f"  🔵 队伍A: {self.team_a_level}  |  🟢 队伍B: {self.team_b_level}",
            ])
        print()

    # ==================== 单局流程 ====================

    def _play_round(self):
        rec = RoundRecord(self.rnd, self.dealer_pid, self.defender_level, self.attacker_level)
        dt = [self.dealer_pid, (self.dealer_pid + 2) % 4]
        at = [(self.dealer_pid + 1) % 4, (self.dealer_pid + 3) % 4]
        rec.dealer_team = dt
        rec.attacker_team = at

        if self.rnd > 1:
            clear()

        self._show_round_header(rec)

        # 发牌
        hands, bottom = self._deal(rec)
        self._show_hands(hands, bottom, rec)

        for stop_count in range(2):
            should_stop = False
            stop_pid = None
            for pid in at:
                if not any(c.rank in SCORE_RANKS for c in hands[pid]):
                    if random.random() < 0.5:
                        should_stop = True
                        stop_pid = pid
                        break
            if should_stop:
                separator()
                label = C.player_label(stop_pid, self.dealer_pid)
                print(f"  {C.BOLD}⏸️ 停级{C.R}")
                print(f"  发起者: {label}")
                print(f"  原因: 手中无分牌（第{stop_count+1}次），系统重新发牌")
                delay(400)
                if not _FAST:
                    wait_prompt("按回车重新发牌")
                hands, bottom = self._deal(rec)
                self._show_hands(hands, bottom, rec)
            else:
                break

        # 定主
        self._determine_trump(rec, hands)
        self._show_trump(rec)

        # 庄家方闷牌 → 自动转亮牌（庄家方不能捡主，只有闲家能捡主）
        if (rec.trump_method == 'concealed'
                and rec.concealed_pid is not None
                and rec.concealed_pid in rec.dealer_team):
            rec.bright_pid = rec.concealed_pid
            rec.bright_card = rec.concealed_card
            rec.trump_suit = rec.concealed_card.suit
            rec.trump_method = 'bright'
            rec.concealed_pid = None
            rec.concealed_card = None

        # 埋底
        self._bury(rec, hands)
        self._show_bury(rec)

        # 捡主
        if rec.concealed_pid is not None:
            self._pick_main(rec, hands)
            self._show_pick_main(rec)
        else:
            delay(200)

        # 出牌
        self._play_tricks(rec, hands, dt, at)

        # 结算
        self._settle(rec)
        self._show_settle(rec)
        return rec

    def _deal(self, rec):
        for attempt in range(100000):
            deck = create_deck()
            random.shuffle(deck)
            hands = [[] for _ in range(4)]
            bottom = []
            for i, card in enumerate(deck):
                (hands[i % 4] if i < 48 else bottom).append(card)
            if not self.deal_requirements or check_deal_requirements(hands, self.deal_requirements):
                rec.initial_hands = {p: list(h) for p, h in enumerate(hands)}
                rec.initial_bottom = list(bottom)
                return hands, bottom
        rec.initial_hands = {p: list(h) for p, h in enumerate(hands)}
        rec.initial_bottom = list(bottom)
        return hands, bottom

    def _determine_trump(self, rec, hands):
        dt, at = rec.dealer_team, rec.attacker_team
        lvl = rec.level

        for pid in range(4):
            lc = [c for c in hands[pid] if c.rank == lvl and c.suit in SUITS]
            if lc and random.random() < 0.25:
                card = random.choice(lc)
                rec.concealed_pid, rec.concealed_card = pid, card
                rec.trump_method = 'concealed'
                return

        for pid in dt:
            lc = [c for c in hands[pid] if c.rank == lvl and c.suit in SUITS]
            if lc and random.random() < 0.5:
                card = random.choice(lc)
                rec.bright_pid, rec.bright_card = pid, card
                rec.trump_suit, rec.trump_method = card.suit, 'bright'
                return

        for pid in at:
            lc = [c for c in hands[pid] if c.rank == lvl and c.suit in SUITS]
            if lc and random.random() < 0.5:
                card = random.choice(lc)
                rec.concealed_pid, rec.concealed_card = pid, card
                rec.trump_method = 'concealed'
                return

        fc = next((c for c in rec.initial_bottom if c.suit in SUITS), None)
        rec.trump_suit = fc.suit if fc else random.choice(SUITS)
        rec.trump_method = 'bottom_card'

    def _bury(self, rec, hands):
        pid = rec.dealer_pid
        bottom = list(rec.initial_bottom)
        bot = Bot(pid, hands[pid], 'dealer', rec.level, rec.trump_suit or '')

        n = random.randint(0, 6)
        buried = bot.select_for_bottom(n)
        temp_bottom = bottom + buried

        score = sum(SCORE_VALUES.get(c.rank, 0) for c in temp_bottom)
        for _ in range(10):
            if score <= 35: break
            sc = [c for c in buried if c.rank in SCORE_RANKS]
            if not sc: break
            buried.remove(sc[0]); bot.hand.append(sc[0])
            temp_bottom = bottom + buried
            score = sum(SCORE_VALUES.get(c.rank, 0) for c in temp_bottom)

        # take_back 数量 = 当前 buried 数量（而非原始 n），确保 new_bottom = 6
        take_back = temp_bottom[:len(buried)]
        new_bottom = temp_bottom[len(buried):]
        assert len(new_bottom) == 6, f"底牌数 {len(new_bottom)} != 6"
        bot.hand.extend(take_back)
        hands[pid] = bot.hand

        rec.buried_cards, rec.bottom_after_bury = list(buried), list(new_bottom)

    def _pick_main(self, rec, hands):
        pid = rec.concealed_pid
        ts = rec.concealed_card.suit
        rec.trump_suit = ts
        bottom = list(rec.bottom_after_bury)
        bot = Bot(pid, hands[pid], 'attacker', rec.level, ts)

        picked = [c for c in bottom if is_main(c, rec.level, ts)]
        if not picked:
            rec.bottom_after_pick = list(bottom)
            return

        rec.picked_from_bottom = list(picked)
        bottom_rem = [c for c in bottom if c not in picked]
        bot.hand.extend(picked)
        discarded = bot.select_for_bottom(len(picked))
        new_bottom = bottom_rem + discarded
        new_bs = sum(SCORE_VALUES.get(c.rank, 0) for c in new_bottom)
        if new_bs > 35:
            rec.bottom_after_pick = list(bottom)
            return
        assert len(new_bottom) == 6
        hands[pid] = list(bot.hand)
        rec.discarded_to_bottom = list(discarded)
        rec.bottom_after_pick = list(new_bottom)

    def _play_tricks(self, rec, hands, dt, at):
        bots = {}
        for pid in range(4):
            side = 'dealer' if pid in dt else 'attacker'
            bots[pid] = Bot(pid, hands[pid], side, rec.level, rec.trump_suit)

        leader = rec.dealer_pid
        t = 0
        max_tricks = 12

        while any(bots[pid].hand for pid in range(4)) and t < max_tricks:
            t += 1
            trick = {'num': t, 'leader': leader, 'played': [], 'winner': None,
                     'winner_side': None, 'score': 0, 'score_cards': [], 'pattern': 'single'}
            lead_suit = None
            played_so_far = []

            for pos in range(4):
                pid = (leader + pos) % 4
                if not bots[pid].hand:
                    trick['played'].append((pid, []))
                    played_so_far.append((pid, []))
                    continue
                if pos == 0:
                    card_list = bots[pid].lead()
                    lead_suit = card_list[0].suit if card_list else None
                else:
                    card_list = bots[pid].follow(lead_suit, played_so_far, is_last_trick=False)
                trick['played'].append((pid, card_list))
                played_so_far.append((pid, card_list))

            if not any(cl for _, cl in trick['played']):
                t -= 1
                continue

            first_cards = played_so_far[0][1] if played_so_far else []
            if first_cards:
                bot0 = bots.get(played_so_far[0][0])
                trick['pattern'] = bot0._detect_pattern(first_cards) if bot0 else 'single'

            best_pid = None
            best_pattern = None
            best_cards = None
            for pid, card_list in trick['played']:
                if not card_list:
                    continue
                p = bots[pid]._detect_pattern(card_list) if pid in bots else 'single'
                if best_pid is None:
                    best_pid, best_pattern, best_cards = pid, p, card_list
                else:
                    cmp = compare_trick_patterns(
                        p, card_list, best_pattern, best_cards,
                        rec.level, rec.trump_suit, lead_suit
                    )
                    if cmp == 1:
                        best_pid, best_pattern, best_cards = pid, p, card_list

            if best_pid is None:
                t -= 1
                continue

            trick['winner'] = best_pid
            trick['winner_side'] = 'dealer' if best_pid in dt else 'attacker'
            trick['winner_pattern'] = best_pattern

            for pid, card_list in trick['played']:
                for card in card_list:
                    if card.rank in SCORE_RANKS:
                        trick['score_cards'].append(card)
                        trick['score'] += SCORE_VALUES[card.rank]

            rec.tricks.append(trick)
            self._show_trick(trick, rec, bots)
            leader = best_pid

        rec.attacker_score = sum(tr['score'] for tr in rec.tricks)
        if rec.tricks:
            lt = rec.tricks[-1]
            rec.last_trick_winner_pid = lt['winner']
            rec.last_trick_winner_side = lt['winner_side']
            for pid, card_list in lt['played']:
                if pid == lt['winner']:
                    rec.last_trick_card = card_list[-1] if card_list else None
                    break

    def _settle(self, rec):
        sc = rec.attacker_score
        is_bottom = rec.last_trick_winner_side == 'attacker'

        if sc == 0:
            rec.result_title = '光头'; rec.base_up_att = 0; rec.final_up_def = 3
        elif sc <= 35:
            rec.result_title = '干受苦'; rec.base_up_att = 0; rec.final_up_def = 1
        elif sc <= 39:
            rec.result_title = '干受苦'; rec.base_up_att = 0; rec.final_up_def = 0
        elif sc <= 45:
            rec.result_title = '上台'; rec.base_up_att = 0; rec.final_up_def = 0
        else:
            rec.base_up_att = min((sc - 50) // 10 + 1, 6)
            rec.final_up_def = 0
            rec.result_title = f"升{rec.base_up_att}级"

        bonus = 0
        if is_bottom:
            bonus = 4 if (rec.last_trick_card and rec.last_trick_card.rank == '大王') else 3
        rec.bonus_up = bonus

        if is_bottom and sc < 40:
            rec.result_title = '干扣底'
            rec.final_up_def = 0; rec.base_up_att = 0; rec.bonus_up = 0

        rec.final_up_att = rec.base_up_att + bonus

        old_def, old_att = self.defender_level, self.attacker_level

        # 记录升级前的队伍等级
        if self.dealer_pid in (0, 2):
            self.team_a_level_before_update = self.team_a_level
            self.team_b_level_before_update = self.team_b_level
        else:
            self.team_b_level_before_update = self.team_b_level
            self.team_a_level_before_update = self.team_a_level

        self.defender_level = level_up(self.defender_level, rec.final_up_def)
        self.attacker_level = level_up(self.attacker_level, rec.final_up_att)

        if self.dealer_pid in (0, 2):
            self.team_a_level = level_up(self.team_a_level, rec.final_up_def)
            self.team_b_level = level_up(self.team_b_level, rec.final_up_att)
            self.team_a_cumulative_steps += rec.final_up_def
            self.team_b_cumulative_steps += rec.final_up_att
        else:
            self.team_b_level = level_up(self.team_b_level, rec.final_up_def)
            self.team_a_level = level_up(self.team_a_level, rec.final_up_att)
            self.team_b_cumulative_steps += rec.final_up_def
            self.team_a_cumulative_steps += rec.final_up_att

        # 过7判定：累计升档数>=13（完整循环），与 gui.py / game.py 一致
        dealer_is_a = self.dealer_pid in (0, 2)
        dealer_steps = self.team_a_cumulative_steps if dealer_is_a else self.team_b_cumulative_steps
        attacker_steps = self.team_b_cumulative_steps if dealer_is_a else self.team_a_cumulative_steps
        dlabel = '队伍A' if dealer_is_a else '队伍B'
        alabel = '队伍B' if dealer_is_a else '队伍A'

        if dealer_steps >= LEVEL_CYCLE_LEN:
            rec.result_title = f'{dlabel}过7🏆'
            rec.round_ended = True
            rec.round_winner = dlabel
            if not self.total_rounds:
                self.game_over = True
                self.winner = f'{dlabel}（庄家方）'
            else:
                self._reset_for_new_round()
            return

        if attacker_steps >= LEVEL_CYCLE_LEN:
            self._reset_for_new_round()
            new_dealer = (self.dealer_pid + 1) % 4 if self.dealer_pid % 2 == 0 else (self.dealer_pid + 3) % 4
            self.dealer_pid = new_dealer
            self.defender_level = '7'
            self.attacker_level = '7'
            rec.result_title = f'{alabel}过7🏆'
            rec.round_ended = True
            rec.round_winner = alabel
            return

        if rec.result_title == '上台' or rec.final_up_att > 0:
            new_dealer = rec.attacker_team[0]
            if new_dealer != self.dealer_pid:
                self.dealer_pid = new_dealer
                self.defender_level, self.attacker_level = self.attacker_level, self.defender_level

    def _reset_for_new_round(self):
        self.team_a_cumulative_steps = 0
        self.team_b_cumulative_steps = 0
        self.defending_team = None
        self.team_a_level = '7'
        self.team_b_level = '7'
        self.defender_level = '7'
        self.attacker_level = '7'


def save_excel(records, game, path):
    wb = Workbook()
    tf = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    tfont = Font(size=14, bold=True, color="FFFFFF")
    hf = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    sf = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    sfont = Font(bold=True, size=11)

    # ====== Sheet1: 游戏总览 ======
    ws = wb.active; ws.title = "游戏总览"
    ws.merge_cells('A1:P1')
    c = ws['A1']; c.value = f"一副牌升级游戏模拟（过7）"; c.font = tfont; c.fill = tf; c.alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:P2')
    total_games = len(records)
    total_r = len(game.round_records) if hasattr(game, 'round_records') else 0
    c = ws['A2']; c.value = f"生成: {datetime.now().strftime('%Y-%m-%d %H:%M')} | 共{total_r}轮{total_games}局 | 胜方: {game.winner or '未完成'}"; c.font = Font(italic=True)

    hdrs = ['轮次','局数','庄家','庄方级(前)','抓方级(前)','本局级',
            '定主方式','亮牌/闷牌','主花色',
            '抓分','扣底','扣底牌',
            '抓方升级','庄方升级','庄方级(后)','抓方级(后)','结果']
    r = 4
    for col, h in enumerate(hdrs, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = hfont; c.fill = hf; c.alignment = Alignment(horizontal='center')

    cur_def, cur_att = '7', '7'
    for i, rec in enumerate(records):
        r = 5 + i
        pre_def, pre_att = cur_def, cur_att

        if rec.trump_method == 'bright':
            tm, td = '亮牌', f"玩{rec.bright_pid+1}亮{rec.bright_card}"
        elif rec.trump_method == 'concealed':
            tm, td = '闷牌', f"玩{rec.concealed_pid+1}闷{rec.concealed_card}"
        else:
            tm, td = '底牌', f"首张{rec.initial_bottom[0] if rec.initial_bottom else '—'}"

        cur_def = level_up(cur_def, rec.final_up_def)
        cur_att = level_up(cur_att, rec.final_up_att)
        ts = rec.trump_suit if rec.trump_suit else '—'
        ib = rec.last_trick_winner_side == 'attacker' if rec.last_trick_winner_side else False

        # 确定轮次
        rnd_num = None
        if hasattr(game, 'round_records'):
            for rr in game.round_records:
                if rr['start_rnd'] <= rec.rnd <= rr['end_rnd']:
                    rnd_num = rr['round']
                    break
            if rnd_num is None:
                rnd_num = game.current_round - 1 if hasattr(game, 'current_round') else '—'

        vals = [f"第{rnd_num}轮" if rnd_num else '—', rec.rnd, f"玩{rec.dealer_pid+1}", pre_def, pre_att, rec.level,
                tm, td, ts, rec.attacker_score,
                '是' if ib else '否', str(rec.last_trick_card or '—'),
                f"+{rec.final_up_att}" if rec.final_up_att else 0,
                f"+{rec.final_up_def}" if rec.final_up_def else 0,
                cur_def, cur_att, rec.result_title]

        for col, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=v); c.alignment = Alignment(horizontal='center')
        # 每轮最后一行高亮
        if hasattr(game, 'round_records'):
            for rr in game.round_records:
                if rec.rnd == rr['end_rnd']:
                    for col in range(1, len(hdrs)+1):
                        ws.cell(row=r, column=col).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for col in range(1, len(hdrs)+1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(hdrs[col-1])*2)

    # ====== Sheet2: 轮次统计 ======
    ws2 = wb.create_sheet("轮次统计")
    ws2.merge_cells('A1:E1')
    c = ws2['A1']; c.value = "轮次统计"; c.font = tfont; c.fill = tf; c.alignment = Alignment(horizontal='center')
    r = 2
    hdrs2 = ['轮次', '起始局', '结束局', '局数', '本轮胜方']
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=r, column=col, value=h)
        c.font = hfont; c.fill = hf; c.alignment = Alignment(horizontal='center')
    r = 3
    for rr in game.round_records:
        vals = [f"第{rr['round']}轮", rr['start_rnd'], rr['end_rnd'], rr['games_count'], rr['winner']]
        for col, v in enumerate(vals, 1):
            ws2.cell(row=r, column=col, value=v).alignment = Alignment(horizontal='center')
        r += 1
    # 汇总行
    if game.round_records:
        total_g = sum(rr['games_count'] for rr in game.round_records)
        for col, v in enumerate(['合计', '', '', total_g, game.winner or '—'], 1):
            c = ws2.cell(row=r, column=col, value=v)
            c.alignment = Alignment(horizontal='center')
            c.font = Font(bold=True)
    for col in range(1, 6):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # ====== Sheet3: 每局详情 ======
    ws3 = wb.create_sheet("每局详情")
    row = 1

    for rec in records:
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws3.cell(row=row, column=1, value=f"═══ 第{rec.rnd}局 ─ {rec.result_title} ═══")
        c.font = tfont; c.fill = tf; row += 1

        for label, val in [
            ("庄家", f"玩家{rec.dealer_pid+1}"),
            ("庄家方级牌", str(rec.defender_level)),
            ("抓分方级牌", str(rec.attacker_level)),
            ("本局级牌", str(rec.level)),
            ("庄家阵营", f"玩家{rec.dealer_team[0]+1}、玩家{rec.dealer_team[1]+1}"),
            ("抓分阵营", f"玩家{rec.attacker_team[0]+1}、玩家{rec.attacker_team[1]+1}"),
            ("定主方式", rec.trump_method or '—'),
            ("主花色", SUIT_CN.get(rec.trump_suit, rec.trump_suit or '—') if rec.trump_suit else '—'),
        ]:
            c = ws3.cell(row=row, column=1, value=label); c.font = sfont
            ws3.cell(row=row, column=3, value=val); row += 1

        if rec.bright_pid is not None:
            c = ws3.cell(row=row, column=1, value="⭐ 亮牌"); c.font = sfont; c.fill = sf; row += 1
            ws3.cell(row=row, column=1, value="亮牌玩家"); ws3.cell(row=row, column=2, value=f"玩家{rec.bright_pid+1}")
            ws3.cell(row=row, column=3, value="亮出的牌"); ws3.cell(row=row, column=4, value=str(rec.bright_card)); row += 1
        if rec.concealed_pid is not None:
            c = ws3.cell(row=row, column=1, value="🃏 闷牌"); c.font = sfont; c.fill = sf; row += 1
            ws3.cell(row=row, column=1, value="闷牌玩家"); ws3.cell(row=row, column=2, value=f"玩家{rec.concealed_pid+1}")
            ws3.cell(row=row, column=3, value="闷的牌"); ws3.cell(row=row, column=4, value=str(rec.concealed_card)); row += 1

        c = ws3.cell(row=row, column=1, value="📋 初始手牌"); c.font = sfont; c.fill = sf; row += 1
        for pid in range(4):
            ws3.cell(row=row, column=1, value=f"玩家{pid+1}")
            ws3.cell(row=row, column=2, value=cards_str_plain(rec.initial_hands.get(pid, [])))
            ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8); row += 1

        c = ws3.cell(row=row, column=1, value="📦 底牌信息"); c.font = sfont; c.fill = sf; row += 1
        ws3.cell(row=row, column=1, value="初始底牌"); ws3.cell(row=row, column=2, value=cards_str_plain(rec.initial_bottom))
        ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8); row += 1

        if rec.buried_cards:
            ws3.cell(row=row, column=1, value="庄家埋入"); ws3.cell(row=row, column=2, value=cards_str_plain(rec.buried_cards))
            ws3.cell(row=row, column=3, value="埋底后底牌"); ws3.cell(row=row, column=4, value=cards_str_plain(rec.bottom_after_bury))
            bs = sum(SCORE_VALUES.get(c.rank,0) for c in rec.bottom_after_bury)
            ws3.cell(row=row, column=5, value=f"分值={bs}")
            ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3); row += 1

        if rec.picked_from_bottom:
            ws3.cell(row=row, column=1, value="拣出主牌"); ws3.cell(row=row, column=2, value=cards_str_plain(rec.picked_from_bottom))
            ws3.cell(row=row, column=3, value="弃回底牌"); ws3.cell(row=row, column=4, value=cards_str_plain(rec.discarded_to_bottom))
            ws3.cell(row=row, column=5, value="捡主后底牌"); ws3.cell(row=row, column=6, value=cards_str_plain(rec.bottom_after_pick)); row += 1

        c = ws3.cell(row=row, column=1, value="📝 操作日志"); c.font = sfont; c.fill = sf; row += 1
        for log in rec.logs:
            ws3.cell(row=row, column=1, value=log)
            ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8); row += 1

        c = ws3.cell(row=row, column=1, value="📊 结算"); c.font = sfont; c.fill = sf; row += 1
        ib = rec.last_trick_winner_side == 'attacker' if rec.last_trick_winner_side else False
        for label, val in [
            ("抓分方得分", str(rec.attacker_score)),
            ("是否扣底", "是" if ib else "否"),
            ("扣底牌", str(rec.last_trick_card) if rec.last_trick_card else "—"),
            ("基础升级 抓方", f"+{rec.base_up_att}" if rec.base_up_att else "0"),
            ("扣底加成", f"+{rec.bonus_up}" if rec.bonus_up else "无"),
            ("总升级 抓方", f"+{rec.final_up_att}"),
            ("总升级 庄方", f"+{rec.final_up_def}"),
            ("结果", rec.result_title),
        ]:
            ws3.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=row, column=3, value=val); row += 1

        row += 2

    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    wb.save(path)
    print(f"✅ Excel: {path}")


class RoundRecord:
    def __init__(self, rnd, dealer_pid, def_lvl, att_lvl):
        self.rnd = rnd
        self.dealer_pid = dealer_pid
        self.defender_level = def_lvl
        self.attacker_level = att_lvl
        self.level = def_lvl
        self.initial_hands = {}
        self.initial_bottom = []
        self.bright_pid = None; self.bright_card = None
        self.concealed_pid = None; self.concealed_card = None
        self.trump_suit = None
        self.trump_method = None
        self.buried_cards = []
        self.bottom_after_bury = []
        self.picked_from_bottom = []
        self.discarded_to_bottom = []
        self.bottom_after_pick = []
        self.tricks = []
        self.logs = []
        self.attacker_score = 0
        self.last_trick_winner_pid = None
        self.last_trick_winner_side = None
        self.last_trick_card = None
        self.base_up_att = 0
        self.bonus_up = 0
        self.final_up_att = 0
        self.final_up_def = 0
        self.result = ''
        self.result_title = ''
        self.dealer_team = []
        self.attacker_team = []
        self.game_over_check = False
        self.round_ended = False
        self.round_winner = None

# ==================== 参数解析 ====================

def _parse_range(val):
    parts = val.split(':')
    lo = int(parts[0])
    hi = int(parts[1]) if len(parts) > 1 else lo
    return (lo, hi)

def parse_args():
    p = argparse.ArgumentParser(description='一副牌升级 · CLI 可视化模拟器')
    p.add_argument('--seed', type=int, default=None, help='随机种子（可复现）')
    p.add_argument('--level', type=str, default='7', help='起始级牌 (默认7)')
    p.add_argument('--max-rounds', type=int, default=200, help='最大局数 (默认200)')
    p.add_argument('--rounds', type=int, default=None, help='总轮数（默认不限制）')
    p.add_argument('--fast', action='store_true', help='快速模式：不等待，自动播放')
    p.add_argument('--single', action='store_true', help='单局模式：只玩一局就结束')
    p.add_argument('--hong', type=str, default=None, metavar='N[:M]', help='双方至少各有玩家拥有N~M个轰')
    p.add_argument('--zha', type=str, default=None, metavar='N[:M]', help='双方至少各有玩家拥有N~M个炸')
    p.add_argument('--510k', type=str, default=None, metavar='N[:M]', dest='pattern_510k', help='双方至少各有玩家拥有N~M个510K')
    return p.parse_args()

def main():
    args = parse_args()
    global _FAST
    _FAST = args.fast

    requirements = {}
    if args.hong:
        requirements['hong'] = _parse_range(args.hong)
    if args.zha:
        requirements['zha'] = _parse_range(args.zha)
    if args.pattern_510k:
        requirements['510k'] = _parse_range(args.pattern_510k)

    max_games = 0 if args.rounds else args.max_rounds
    max_rds = 1 if args.single else args.max_rounds

    game = CLIGame(
        seed=args.seed,
        start_level=args.level,
        max_rounds=max_rds,
        total_rounds=args.rounds,
        max_games=max_games,
        deal_requirements=requirements or None,
    )
    game.run()

    # 打印轮次统计
    if game.round_records:
        print(f"\n📊 共 {len(game.round_records)} 轮 {len(game.records)} 局 | 胜方: {game.winner or '未完成'}")
        print("\n📊 轮次统计:")
        for rr in game.round_records:
            print(f"  第{rr['round']}轮: 局{rr['start_rnd']}~{rr['end_rnd']} 共{rr['games_count']}局 | 胜方: {rr['winner']}")

        # 保存 Excel
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_dir = os.path.join(os.path.dirname(__file__), 'output')
        os.makedirs(output_dir, exist_ok=True)
        path = os.path.join(output_dir, f'一副牌升级游戏模拟_{ts}.xlsx')
        save_excel(game.records, game, path)

if __name__ == '__main__':
    main()
